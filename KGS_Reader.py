from collections import defaultdict
import datetime
import json
import os
import re
import shutil
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText

# Попытка импортировать необходимые библиотеки с обработкой ошибок
try:
    import fitz  # PyMuPDF
    PDF_SUPPORTED = True
except ImportError:
    fitz = None
    PDF_SUPPORTED = False
    print("Внимание: Библиотека PyMuPDF (fitz) не установлена. Обработка PDF будет недоступна.")

try:
    import pytesseract
    from PIL import Image, ImageEnhance, ImageFilter
    OCR_SUPPORTED = True
except ImportError:
    pytesseract = None
    Image = None
    ImageEnhance = None
    ImageFilter = None
    OCR_SUPPORTED = False
    print("Внимание: Библиотеки pytesseract и/или PIL не установлены. OCR будет недоступен.")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    EXCEL_SUPPORTED = True
except ImportError:
    Workbook = None
    load_workbook = None
    EXCEL_SUPPORTED = False
    print("Внимание: Библиотека openpyxl не установлена. Работа с Excel будет недоступна.")


# ======================= БИЗНЕС-ЛОГИКА =======================

class PDFProcessor:
    """
    Класс, который занимается ТОЛЬКО обработкой данных, без прямых обращений к GUI.
    Все сообщения пользователю идут через log_callback.
    """

    def __init__(self, log_callback=None):
        self.log_callback = log_callback or print
        self.log_file_path = "application_log.txt"
        self.output_excel_path = ""
        self.problem_files = []
        self.field_stats = defaultdict(int)
        self.import_points = False
        self.points_folder = ""
        self.debug_mode = False
        self.ignore_excel = False
        self.sort_points_by_comm = False  # раскладывать каталоги по типам
        self.ocr_ready = False

        # Конфиг типов коммуникаций
        self.comm_types_config_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "comm_types.json"
        )
        self.default_comm_types = self._build_default_comm_types()
        self.comm_types = []  # список словарей: {"name": str, "enabled": bool}
        self.load_comm_types()

        self.setup_tesseract()

    # ----- Конфиг типов коммуникаций -----

    def _build_default_comm_types(self):
        """Список типов коммуникаций по умолчанию (как было в коде)."""
        types = [
            # Связь и телекоммуникации
            "Кабель связи", "Тел канализация", "Кабельная канализация", "ВОЛС", "КТВ",
            # Электроснабжение
            "Эл кабель", "Кабель техн. и очаг. заземл", "Контур заземл", "Кабель но",
            # Водоснабжение и водоотведение
            "Водосток", "Вод-д", "Трубопровод", "Канализация хоз-быт", "ЛОС", "Дренаж",
            # Теплоснабжение и вентиляция
            "Воздухопровод", "Вент. ветки", "Теплотрасса",
            # Инфраструктурные системы
            "Коллектор",
            # Инженерные системы
            "Газопровод", "Нефтепровод", "Продуктопровод",
            # Специальные системы
            "СОУЭ", "СКУД",
            # Дополнительные типы
            "Канализация",           # общая
            "Нап канализация",       # напорная
            "Сам канализация",       # самотечная
            "Водовыпуск",            # спец. тип
            "Кабель защ",            # кабель защиты
            "Газ",                   # как в папке
            # "Теплотрасса" — уже есть выше
        ]
        # убираем дубликаты, сохраняя порядок
        seen = set()
        unique = []
        for t in types:
            if t not in seen:
                seen.add(t)
                unique.append(t)
        return unique

    def load_comm_types(self):
        """Загрузить типы коммуникаций из JSON или взять значения по умолчанию."""
        if os.path.exists(self.comm_types_config_path):
            try:
                with open(self.comm_types_config_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                types = data.get("types", [])
                cleaned = []
                names_seen = set()
                for t in types:
                    name = str(t.get("name", "")).strip()
                    enabled = bool(t.get("enabled", True))
                    if not name or name in names_seen:
                        continue
                    names_seen.add(name)
                    cleaned.append({"name": name, "enabled": enabled})
                # Если файл есть, но типов нет — fallback на дефолт
                if cleaned:
                    self.comm_types = cleaned
                else:
                    self.reset_comm_types_to_defaults(save=False)
            except Exception as e:
                self.log_message(f"Не удалось загрузить comm_types.json: {e}. Использую значения по умолчанию.")
                self.reset_comm_types_to_defaults(save=False)
        else:
            self.reset_comm_types_to_defaults(save=False)

    def save_comm_types(self):
        """Сохранить текущий список типов коммуникаций."""
        try:
            data = {"types": self.comm_types}
            with open(self.comm_types_config_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self.log_message(f"Типы коммуникаций сохранены: {self.comm_types_config_path}")
        except Exception as e:
            self.log_message(f"Ошибка сохранения comm_types.json: {e}")

    def reset_comm_types_to_defaults(self, save=True):
        """Сброс к значениям по умолчанию: все дефолтные типы включены."""
        self.comm_types = [{"name": name, "enabled": True} for name in self.default_comm_types]
        if save:
            self.save_comm_types()

    def get_comm_types(self):
        """Для GUI: получить текущий список типов (имя + флаг enabled)."""
        return [dict(name=t["name"], enabled=bool(t.get("enabled", True))) for t in self.comm_types]

    def update_comm_types(self, types_list):
        """
        Обновить список типов из GUI.
        types_list: [{"name": str, "enabled": bool}, ...]
        """
        cleaned = []
        names_seen = set()
        for t in types_list:
            name = str(t.get("name", "")).strip()
            if not name or name in names_seen:
                continue
            enabled = bool(t.get("enabled", True))
            names_seen.add(name)
            cleaned.append({"name": name, "enabled": enabled})
        # Чтобы не потерять новые дефолтные типы (если код обновился),
        # можно по желанию слить их. Пока считаем, что конфиг — источник истины.
        self.comm_types = cleaned
        self.save_comm_types()

    def get_allowed_comm_types(self):
        """Список типов, которые сейчас считаются 'ожидаемыми' (enabled=True)."""
        if not self.comm_types:
            # на всякий случай
            return self.default_comm_types
        return [t["name"] for t in self.comm_types if t.get("enabled", True)]

    # ----- Логирование -----

    def log_message(self, message):
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{ts}] {message}\n"
        # в GUI или в консоль
        self.log_callback(message)
        # файл
        try:
            if self.log_file_path:
                with open(self.log_file_path, "a", encoding="utf-8") as f:
                    f.write(line)
        except Exception as e:
            # если лог не запишется — не критично, просто сообщаем в GUI/консоль
            self.log_callback(f"ОШИБКА логирования: {e}")

    # ----- OCR / Tesseract -----
 
    def setup_tesseract(self):
        """Проверка наличия Tesseract. Никаких messagebox — только лог."""
        if not OCR_SUPPORTED or pytesseract is None:
            self.log_message("OCR не поддерживается: библиотеки не установлены")
            self.ocr_ready = False
            return
        try:
            def possible_tess_paths():
                # 1) рядом с exe/скриптом (portable сборка)
                exe_dir = os.path.dirname(getattr(sys, "_MEIPASS", sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)))
                local_candidates = [
                    os.path.join(exe_dir, "tesseract", "tesseract.exe"),
                    os.path.join(exe_dir, "tesseract.exe"),
                ]
                # 2) стандартные инсталляции
                win_candidates = [
                    r'C:\Program Files\Tesseract-OCR\tesseract.exe',
                    r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
                    os.path.join(os.environ.get('LOCALAPPDATA', ''), 'Tesseract-OCR', 'tesseract.exe'),
                    os.path.join(os.path.expanduser('~'), 'AppData', 'Local', 'Tesseract-OCR', 'tesseract.exe')
                ]
                return local_candidates + win_candidates

            if sys.platform == 'win32':
                paths = possible_tess_paths()
                for p in paths:
                    if os.path.exists(p):
                        pytesseract.pytesseract.tesseract_cmd = p
                        break
            pytesseract.get_tesseract_version()
            self.log_message("Tesseract OCR: ок")
            self.ocr_ready = True
        except Exception as e:
            self.log_message(f"Tesseract не найден/не работает: {e}")
            self.ocr_ready = False

    def enhance_image(self, image):
        if not OCR_SUPPORTED or ImageEnhance is None or ImageFilter is None:
            return image
        image = ImageEnhance.Contrast(image).enhance(2.0)
        image = ImageEnhance.Sharpness(image).enhance(2.0)
        image = image.convert('L')
        image = image.filter(ImageFilter.MedianFilter(size=3))
        image = image.point(lambda x: 0 if x < 140 else 255)
        return image

    def extract_text_with_ocr(self, page):
        if not (OCR_SUPPORTED and PDF_SUPPORTED and fitz and pytesseract and Image and self.ocr_ready):
            self.log_message("OCR не поддерживается или Tesseract не готов")
            return ""
        try:
            matrix = fitz.Matrix(4, 4)
            pix = page.get_pixmap(matrix=matrix, dpi=300)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img = self.enhance_image(img)
            return pytesseract.image_to_string(img, config=r'--oem 3 --psm 3 -l rus+eng --dpi 300')
        except Exception as e:
            self.log_message(f"OCR ошибка: {e}")
            return ""

    # ----- Обработка PDF -----

    def process_pdf(self, file_path):
        if not PDF_SUPPORTED or fitz is None:
            self.log_message("Обработка PDF не поддерживается: библиотека PyMuPDF не установлена")
            return None
        try:
            doc = fitz.open(file_path)
            full_text = ""
            for i, page in enumerate(doc):
                text = page.get_text("text")
                if not text or len(text.strip()) < 50:
                    self.log_message(f"Стр.{i+1}: OCR")
                    text = self.extract_text_with_ocr(page)
                full_text += text + "\n"
            doc.close()
            return full_text
        except Exception as e:
            self.log_message(f"Ошибка PDF {os.path.basename(file_path)}: {e}")
            return None

    # ----- Извлечение КГС -----

    def extract_kgc_number(self, text):
        patterns = [
            r"№ КГС:\s*(\d{2,5}[-\/]\d{2,5})",
            r"№ КГС:\s*([A-ZА-Я]?\d+[A-ZА-Я]?)",
            r"(?:КГС|№|N)\s*[:\-]?\s*(\d{2,5}[-\/]\d{2,5})",
            r"(?:КГС|№|N)\s*[:\-]?\s*([A-ZА-Я]?\d+[A-ZА-Я]?)",
            r"КГС\s*([^\n,;]+)",
            r"\b(\d{2,5}-\d{2,5})\b",
            r"\b(\d{5}-\d{2})\b"
        ]
        for p in patterns:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                kgs = (m.group(1) if m.groups() else m.group(0)).strip()
                kgs = re.sub(r'[^\dА-ЯA-Z\-/]', '', kgs)
                if len(kgs) >= 4:
                    return kgs
        return None

    # ----- Similarity и нормализация типа коммуникации -----

    def similarity(self, a, b):
        a, b = a.lower(), b.lower()
        if a in b or b in a:
            return 0.9

        a_words = set(a.split())
        b_words = set(b.split())
        if not a_words or not b_words:
            return 0.0
        common_words = a_words & b_words
        union_words = a_words | b_words
        similarity_score = len(common_words) / len(union_words)

        if len(a) > 3 and len(b) > 3:
            a_short = ''.join([w[0] for w in a.split() if w])
            b_short = ''.join([w[0] for w in b.split() if w])
            if a_short and b_short and a_short == b_short:
                similarity_score = max(similarity_score, 0.7)

        return similarity_score

    def normalize_communication_type(self, text, allowed_communications):
        if not text:
            return None

        text = text.strip()

        normalization_map = {
            # Связь и телекоммуникации
            r'кабел[ьи]?\s*связ[и]?': 'Кабель связи',
            r'тел[е]?\s*канализац[ия]{2,4}': 'Тел канализация',
            r'кабел[ьи]?\s*канализац[ия]{2,4}': 'Кабельная канализация',
            r'волоконно-?\s*оптическ[ая]?\s*лини[яи]?\s*связ[и]?': 'ВОЛС',
            r'кабел[ьное]?\s*телевидени[е]?': 'КТВ',

            # Электроснабжение
            r'эл[ек]?\s*кабел[ья]?': 'Эл кабель',
            r'кабел[ь]?\s*техн[\.]?\s*и\s*очаг[\.]?\s*заземл[ения]?': 'Кабель техн. и очаг. заземл',
            r'контур\s*заземл[ения]?': 'Контур заземл',
            r'кабел[ь]?\s*н[оo0]': 'Кабель но',
            r'наружн[ое]?\s*освещени[е]': 'Кабель но',

            # Водоснабжение и водоотведение
            r'ливнев[ая]?\s*канализац[ия]{2,4}': 'Водосток',
            r'вод-?д': 'Вод-д',
            r'водопровод': 'Вод-д',
            r'трубопровод': 'Трубопровод',
            r'канализац[ия]{2,4}\s*хоз-?быт': 'Канализация хоз-быт',
            r'хоз-?бытов[ая]?\s*канализац[ия]{2,4}': 'Канализация хоз-быт',
            r'лос\b': 'ЛОС',
            r'локальн[ые]?\s*очистн[ые]?\s*сооружени[я]': 'ЛОС',
            r'дренаж': 'Дренаж',

            # Теплоснабжение и вентиляция
            r'воздухопровод': 'Воздухопровод',
            r'вент[\.]?\s*ветк[и]?': 'Вент. ветки',
            r'вентиляционн[ые]?\s*ветк[и]?': 'Вент. ветки',
            r'теплотрасса': 'Теплотрасса',
            r'теплов[ые]?\s*сет[и]?': 'Теплотрасса',

            # Инфраструктурные системы
            r'коллектор': 'Коллектор',

            # Инженерные системы
            r'газопровод': 'Газопровод',
            r'нефтепровод': 'Нефтепровод',
            r'продуктопровод': 'Продуктопровод',

            # Специальные системы
            r'соуэ\b': 'СОУЭ',
            r'систем[аы]?\s*оповещени[я]?\s*и\s*управлен[ие]?\s*эвакуаци[ей]': 'СОУЭ',
            r'скуд\b': 'СКУД',
            r'систем[аы]?\s*контрол[я]?\s*управлени[я]?\s*доступом': 'СКУД',

            # Канализация: общая, напорная, самотечная
            r'\bканализац(ия)?\b': 'Канализация',
            r'нап[оо]рн[аяые]?\s*канализац[ия]{2,4}|нап\.?\s*канализац': 'Нап канализация',
            r'сам[о]?течн[аяые]?\s*канализац[ия]{2,4}|сам\.?\s*канализац': 'Сам канализация',

            # Водовыпуск
            r'водо[вв]ыпуск': 'Водовыпуск',

            # Кабель защиты
            r'кабел[ьи]?\s*защ': 'Кабель защ',

            # Газ
            r'\bгаз\b': 'Газ',
            r'газопровод': 'Газопровод',

            # Теплоснабжение (общий паттерн)
            r'тепл(о|\.|\b)': 'Теплотрасса',
        }

        # Сначала пробуем нормализацию по паттернам
        for pattern, standard in normalization_map.items():
            if re.search(pattern, text, re.IGNORECASE):
                if standard in allowed_communications:
                    return standard

        # Если не совпало с явным паттерном — пробуем прямое совпадение
        for comm_type in allowed_communications:
            if self.similarity(text.lower(), comm_type.lower()) > 0.9:
                return comm_type

        return None

    def find_best_communication_match(self, text, allowed_communications):
        best_match = None
        best_score = 0.6  # повышенный порог

        if not allowed_communications:
            return None

        words_to_check = []
        lines = text.split('\n')

        for line in lines:
            words = line.strip().split()
            for i in range(len(words) - 1):
                phrase = ' '.join(words[i:i+2])
                if len(phrase) > 5:
                    words_to_check.append(phrase)
                if i < len(words) - 2:
                    phrase = ' '.join(words[i:i+3])
                    if len(phrase) > 8:
                        words_to_check.append(phrase)

        for phrase in words_to_check:
            for comm_type in allowed_communications:
                score = self.similarity(phrase.lower(), comm_type.lower())
                if score > best_score:
                    best_score = score
                    best_match = comm_type

        if best_match:
            self.log_message(f"Определен тип коммуникации: '{best_match}' (схожесть: {best_score:.1%})")
            return best_match

        return None

    # ----- Извлечение полей -----

    def extract_field(self, text, field_name, patterns):
        if field_name == "Тип коммуникации":
            allowed_communications = self.get_allowed_comm_types()

            # Паттерны для прямого поиска
            patterns += [
                # Связь и телекоммуникации
                r"Кабель\s*связи\b", r"Тел\s*канализац[ия]{2,4}\b", r"Кабельная\s*канализац[ия]{2,4}\b",
                r"ВОЛС\b", r"КТВ\b",
                # Электроснабжение
                r"Эл\s*кабель\b", r"Кабель\s*техн\.?\s*и\s*очаг\.?\s*заземл\b", r"Контур\s*заземл\b",
                r"Кабель\s*н[оo0]\b",
                # Водоснабжение и водоотведение
                r"Водосток\b", r"Вод-?д\b", r"Трубопровод\b", r"Канализац[ия]{2,4}\s*хоз-?быт\b",
                r"ЛОС\b", r"Дренаж\b",
                # Теплоснабжение и вентиляция
                r"Воздухопровод\b", r"Вент\.?\s*ветки\b", r"Теплотрасса\b",
                # Инфраструктурные системы
                r"Коллектор\b",
                # Инженерные системы
                r"Газопровод\b", r"Нефтепровод\b", r"Продуктопровод\b",
                # Специальные системы
                r"СОУЭ\b", r"СКУД\b",
                r'\bКанализац(ия)?\b', r'Нап\s*канализац', r'Сам\s*канализац',
                r'Водовыпуск\b', r'Кабель\s*защ\b', r'\bГаз\b', r'\bТепл\b'
            ]

            # Сначала прямые паттерны
            for p in patterns:
                m = re.search(p, text, re.IGNORECASE)
                if m:
                    found_text = (m.group(1) if (m.groups() and m.group(1)) else m.group(0)).strip()
                    normalized = self.normalize_communication_type(found_text, allowed_communications)
                    if normalized:
                        return normalized

            # Если прямых совпадений нет — пытаемся интеллектуальное сопоставление
            best_match = self.find_best_communication_match(text, allowed_communications)
            if best_match:
                return best_match

            return None

        # Остальные поля — старая логика
        for p in patterns:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                val = (m.group(1) if (m.groups() and m.group(1)) else m.group(0)).strip()
                if field_name == "Номер договора":
                    val = re.sub(r'[^0-9A-Za-zА-Яа-я\/\-]', '', val)
                    m2 = re.search(
                        r'(\b\d+\/[A-ZА-Я]+\/[\wА-Яа-я]+\-?\d+\/\d+\b|\b\d+\/[A-ZА-Я]+\-?\d+\/\d+\b|\b\d+\/\d+\-?\d+\b|\b\d+\/[A-ZА-Я]+\-\d+\b|\b\d{1,2}\/\d{5}\-?\d{1,2}\b|\b\d{1,2}\/\d{5}\b|\b[A-ZА-Я]+\-\d+\/\d+\b|\b\d{1,5}[A-ZА-Я]*[-\/]\d{1,5}\b)',
                        val, re.IGNORECASE
                    )
                    return m2.group(1) if m2 else val
                return val
        return None

    def extract_data(self, text):
        data = {}

        data["Тип коммуникации"] = self.extract_field(text, "Тип коммуникации", [
            r"Вид\s*коммуникации/здания,\s*сооружения:\s*([^\n]+)",
            r"Вид\s*коммуникации[^\n:]*[:\s]*([^\n]+)",
            r"Тип\s*коммуникации[^\n:]*[:\s]*([^\n]+)",
            r"Коммуникац[ияи][^\n:]*[:\s]*([^\n]+)",
        ])

        data["Номер договора"] = self.extract_field(text, "Номер договора", [
            r"№\s*договора\s*\(?соглашения\)?\s*на\s*проведение\s*работ[^\n:]*[:\s]*([^\n,;]+)",
            r"№\s*договора[^\n:]*[:\s]*([^\n,;]+)",
            r"Договор\s*№\s*(\S+)",
            r"№\s*контракта[^\n:]*[:\s]*(\S+)"
        ])
        data["КГС"] = self.extract_kgc_number(text)
        data["Дата съемки"] = self.extract_field(text, "Дата съемки", [
            r"Дата\s*съ[её]мки\s*[:\s]*([0-9]{2}\.[0-9]{2}\.[0-9]{4})",
            r"Съемка\s*от\s*([0-9]{2}\.[0-9]{2}\.[0-9]{4})",
            r"\b\d{2}\.\d{2}\.\d{4}\b"
        ])

        return data

    # ----- Вспомогательное: имена файлов -----

    def sanitize_filename(self, s):
        if not s:
            return "UNKNOWN_KGS"
        s = re.sub(r'[\\/*?:"<>|]', '_', str(s).strip())
        return re.sub(r'\s+', '_', s)

    def _comm_subfolder(self, base_folder, comm_type):
        name = self.sanitize_filename(comm_type or "Без_типа")
        return os.path.join(base_folder, name)

    # ----- Каталог координат -----
    # (эта часть полностью как у тебя, только без messagebox; я её не трогаю
    #   по структуре, чтобы не ломать функционал)

    def extract_and_save_coordinate_table(self, document_text, kgs, out_folder, src_pdf):
        """
        Возврат: (points_status, points_count_str, count_found, max_id)
        + Грубые проверки/правки:
          - восстановление пропущенного минуса у X/Y (по большинству знаков)
          - правка высоты вида 17093 -> 170.93
          - фиксация неполных/сомнительных строк в отдельный отчет
        """
        if not kgs:
            return "Нет КГС", "0/0", 0, 0

        fname = os.path.join(out_folder, f"{self.sanitize_filename(kgs)}.txt")
        issues_name = os.path.join(out_folder, f"{self.sanitize_filename(kgs)}_issues.txt")

        start_keys = [
            r"каталог\s+(?:исполнительных\s+|фактических\s+)?координат",
            r"ведомость\s+(?:исполнительных\s+|фактических\s+)?координат",
            r"координаты\s+точек", r"координаты\s+пунктов",
            r"№\s*точки.*X.*[YУ].*[HН]?", r"n/n\s*по\s*съемке\s*[xх]\s*,\s*[мМm]\s*[yу]\s*,\s*[мМm]\s*[hн]\s*,\s*[мМm]"
        ]
        row_re = re.compile(
            r"^\s*(\d+)(?:\s+\d+)?\s+([-–—−]?\d{1,3}(?:\s*\d{3})*[.,]\d{1,3})\s+([-–—−]?\d{1,3}(?:\s*\d{3})*[.,]\d{1,3})(?:\s+([-–—−]?\d{1,3}(?:\s*\d{3})*[.,]\d{1,3}))?(?:\s+(.*))?$",
            re.IGNORECASE
        )
        num_token_re = re.compile(r"[-–—−]?\d+(?:\s?\d{3})*(?:[.,]\d+)?")

        def clean_num_string(v: str) -> str:
            if not v:
                return ""
            s = str(v).strip()
            repl = {'O': '0', 'О': '0', 'I': '1', 'L': '1', 'Е': '1', 'Z': '2', 'З': '3', 'S': '5', 'Б': '6',
                    'B': '8', 'В': '8', '°': '0', '=': '-', '−': '-', '–': '-', '—': '-', '_': '-'}
            for a, b in repl.items():
                s = s.replace(a, b)
            s = re.sub(r'\s+', '', s).replace(',', '.')
            if s.count('-') > 1:
                s = ('-' + s.replace('-', '')) if s.startswith('-') else s.replace('-', '')
            if s.count('.') > 1:
                p = s.split('.')
                s = p[0] + '.' + ''.join(p[1:])
            s = re.sub(r'[^\d\.-]', '', s)
            return s if re.search(r'\d', s) else ""

        def as_float(s):
            try:
                return float(s)
            except Exception:
                return None

        def fuzzy_parse(line: str):
            m_id = re.match(r"\s*(\d+)(?:\s+\d+)?", line)
            if not m_id:
                return None
            pid = m_id.group(1)
            rest = line[m_id.end():]
            nums = [clean_num_string(t.group(0)) for t in num_token_re.finditer(rest)]
            nums = [n for n in nums if n]

            def drop_leading_extra_id(seq):
                """
                Убираем второй номер точки, если он прилип перед координатами
                (пример: '22 23 -12929.73 -1701.87' -> координаты начинаются после 23).
                """
                if len(seq) < 3:
                    return seq
                first, second = seq[0], seq[1]
                looks_like_id = re.fullmatch(r"-?\d{1,4}", first or "")
                looks_like_coord = ('.' in second) or (second and abs(float(second)) > 10000)
                if looks_like_id and looks_like_coord:
                    return seq[1:]
                return seq

            nums = drop_leading_extra_id(nums)
            x = y = h = ""
            if len(nums) >= 2:
                x, y = nums[0], nums[1]
                if len(nums) >= 3:
                    h = nums[2]
            desc = ""
            if nums:
                last = None
                for m in num_token_re.finditer(rest):
                    last = m
                if last:
                    desc = rest[last.end():].strip()
            else:
                desc = rest.strip()

            return pid, x, y, h, desc

        lines = document_text.splitlines()
        raw_rows = []
        parsing = False
        skip = 0
        MAX_SKIP = 10
        issues = []
        max_id = 0
        table_header_found = False

        for line in lines:
            t = line.strip()
            if not t:
                if parsing:
                    skip += 1
                    if skip >= MAX_SKIP and raw_rows:
                        break
                continue
            if not parsing:
                if any(re.search(k, t, re.IGNORECASE) for k in start_keys):
                    parsing, skip = True, 0
                continue

            if parsing and not table_header_found:
                table_header_patterns = [
                    r"n/n\s*по\s*съемке\s*[xх]\s*,\s*[мmМ]\s*[yу]\s*,\s*[мmМ]\s*[hн]\s*,\s*[мmМ]?",
                    r"n/n\s*по\s*съемке\s*[xх]\s*,?\s*[мmМ]?\s*[yу]\s*,?\s*[мmМ]?\s*[hн]\s*,?\s*[мmМ]?",
                    r"№\s*точки.*[xхyуhн]",
                    r"№\s*точки.*координаты"
                ]
                if any(re.search(pattern, t, re.IGNORECASE) for pattern in table_header_patterns):
                    table_header_found = True
                    skip = 0
                continue

            m = row_re.match(t)
            if m:
                skip = 0
                pid, x, y, h, d = m.groups()
                x, y, h = clean_num_string(x), clean_num_string(y), clean_num_string(h)
            else:
                skip += 1
                parsed = fuzzy_parse(t)
                if not parsed:
                    if raw_rows and skip >= MAX_SKIP:
                        break
                    else:
                        continue
                pid, x, y, h, d = parsed
                x, y, h = clean_num_string(x), clean_num_string(y), clean_num_string(h)

            if pid and pid.isdigit():
                max_id = max(max_id, int(pid))
            raw_rows.append({"pid": pid, "x": x, "y": y, "h": h, "d": (d or "").strip(), "line": t})

        if not raw_rows:
            self.log_message(f"Каталог координат не найден ({src_pdf})")
            return "Нет точек", "0/0", 0, 0

        xs = [as_float(r["x"]) for r in raw_rows if r["x"]]
        ys = [as_float(r["y"]) for r in raw_rows if r["y"]]
        hs = [as_float(r["h"]) for r in raw_rows if r["h"]]

        def majority_negative(values):
            vs = [v for v in values if v is not None]
            if not vs:
                return False
            neg = sum(1 for v in vs if v < 0)
            pos = sum(1 for v in vs if v > 0)
            return neg > pos

        x_should_be_negative = majority_negative(xs)
        y_should_be_negative = majority_negative(ys)

        fixed_rows = []
        for r in raw_rows:
            pid, x_s, y_s, h_s, d = r["pid"], r["x"], r["y"], r["h"], r["d"]
            x, y, h = as_float(x_s), as_float(y_s), as_float(h_s)
            notes = []

            if x is None or y is None:
                notes.append("неполная строка (нет X или Y)")

            def maybe_fix_sign(val, s, should_neg, label):
                if val is None:
                    return val, s, False
                if should_neg and val > 0 and abs(val) > 500:
                    s_fixed = "-" + s if not s.startswith("-") else s
                    notes.append(f"возможен потерянный минус у {label} -> исправил")
                    return -val, s_fixed, True
                return val, s, False

            x, x_s, _ = maybe_fix_sign(x, x_s, x_should_be_negative, "X")
            y, y_s, _ = maybe_fix_sign(y, y_s, y_should_be_negative, "Y")

            if h is not None and h > 500:
                if re.fullmatch(r"\d{4,6}", h_s or ""):
                    h_s_fixed = h_s[:-2] + "." + h_s[-2:]
                    h_fixed = as_float(h_s_fixed)
                    if h_fixed is not None and 0 < h_fixed < 500:
                        notes.append(f"высота без точки ({h_s}) -> {h_s_fixed}")
                        h, h_s = h_fixed, h_s_fixed

            if ("неполная строка" in " ".join(notes)) or x is None or y is None:
                issues.append(f"{pid}\t{x_s or ''}\t{y_s or ''}\t{h_s or ''}\t{d}    <-- ПРОБЛЕМА: {', '.join(notes) or 'не удалось распарсить'}")
            fixed_rows.append((pid, x_s or "", y_s or "", h_s or "", d))

        # Детекция выбросов по диапазону координат
        xs_nonnull = [as_float(r.get("x")) for r in raw_rows if r.get("x")]
        ys_nonnull = [as_float(r.get("y")) for r in raw_rows if r.get("y")]
        def calc_threshold(vals):
            vals = [abs(v) for v in vals if v is not None]
            if not vals:
                return None
            median_val = sorted(vals)[len(vals)//2]
            # разумный максимум: либо 100_000, либо 10 * медиана
            return max(100_000, median_val * 10)

        x_threshold = calc_threshold(xs_nonnull)
        y_threshold = calc_threshold(ys_nonnull)

        os.makedirs(out_folder, exist_ok=True)
        with open(fname, 'w', encoding='utf-8') as f:
            for pid, x_s, y_s, h_s, d in fixed_rows:
                if x_s and y_s:
                    f.write(f"{pid}\t{x_s}\t{y_s}\t{h_s}\t{d}\n")

                try:
                    xv = float(x_s) if x_s else None
                    yv = float(y_s) if y_s else None
                    outlier = (
                        (x_threshold is not None and xv is not None and abs(xv) > x_threshold) or
                        (y_threshold is not None and yv is not None and abs(yv) > y_threshold)
                    )
                    if outlier:
                        issues.append(f"{pid}\t{x_s}\t{y_s}\t{h_s}\t{d}    <-- ПРОБЛЕМА: возможный выброс координаты")
                except Exception:
                    pass

        if issues:
            with open(issues_name, 'w', encoding='utf-8') as f:
                f.write("Строки с подозрениями/ошибками (после грубых автоправок):\n")
                for it in issues:
                    f.write(it + "\n")
            self.log_message(f"⚠ Обнаружены проблемные строки: {os.path.basename(issues_name)}")

        cnt = sum(1 for pid, x_s, y_s, h_s, d in fixed_rows if x_s and y_s)
        count_str = f"{cnt}/{max_id if max_id else cnt}"
        self.log_message(f"Каталог: {os.path.basename(fname)} | Точек: {count_str}")
        return "Точки сохранены", count_str, cnt, max_id

    # ----- Excel -----

    def adjust_columns(self, ws):
        if not EXCEL_SUPPORTED:
            return
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            width = min(max((len(str(c.value)) for c in col if c.value), default=0) + 2, 50)
            ws.column_dimensions[get_column_letter(col[0].column)].width = width

    def apply_standard_excel_style(self, ws, headers):
        if not EXCEL_SUPPORTED:
            return

        try:
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            body_font = Font(size=11, name="Calibri")
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(style="thin", color="9E9E9E")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            max_row = ws.max_row
            max_col = ws.max_column
            if max_row < 1 or max_col < 1:
                return

            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = border
            ws.row_dimensions[1].height = 22

            for r in range(2, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = body_font
                    cell.alignment = center
                    cell.border = border

            try:
                date_col_idx = headers.index("Дата съемки") + 1
                for r in range(2, max_row + 1):
                    ws.cell(row=r, column=date_col_idx).number_format = "DD.MM.YYYY"
            except ValueError:
                pass

            ws.freeze_panes = "A2"
            if max_row > 1:
                ws.auto_filter.ref = ws.dimensions

        except Exception as e:
            self.log_message(f"Не удалось применить стиль Excel: {e}")

    # ----- Сводная статистика -----

    def analyze_results(self, total):
        self.log_message("--- Сводка ---")
        if not total:
            self.log_message("Нет обработанных файлов.")
            return
        for f in ["Тип коммуникации", "Номер договора", "КГС", "Дата съемки"]:
            c = self.field_stats.get(f, 0)
            self.log_message(f"{f}: {c}/{total} ({(c/total*100):.1f}%)")

    # ----- Основная обработка выбранных файлов -----

    def process_selected_files(self, folder_path, selected_filenames, target_move_folder=None):
        """
        ВАЖНО: здесь нет messagebox — только лог и возвращаемые значения.
        Проверку существования папки делает GUI до вызова.
        """
        if not os.path.exists(folder_path):
            self.log_message(f"Папка не существует: {folder_path}")
            return None

        output_path = os.path.join(folder_path, "Реестр_геодезических_съемок.xlsx")
        self.output_excel_path = output_path
        self.log_file_path = os.path.join(folder_path, "application_log.txt")

        headers = ["Файл", "Тип коммуникации", "Номер договора", "КГС", "Дата съемки", "Количество точек", "Статус", "Точки"]

        wb = ws = None
        existing = set()
        processed = 0
        moved = 0
        excel_created_or_updated = False

        if not self.ignore_excel:
            if not EXCEL_SUPPORTED:
                self.log_message("Работа с Excel не поддерживается: библиотека openpyxl не установлена")
            else:
                if os.path.exists(output_path):
                    try:
                        wb = load_workbook(output_path) if load_workbook else None
                        if wb is not None:
                            ws = wb.active
                            if ws is not None:
                                current_headers = [cell.value for cell in ws[1] if cell.value]
                                if current_headers != headers:
                                    self.log_message("Обновляю структуру Excel файла...")
                                    new_ws = wb.create_sheet("Геодезия_новый")
                                    new_ws.append(headers)

                                    for row in ws.iter_rows(min_row=2, values_only=True):
                                        if row and row[0]:
                                            adjusted_row = list(row)[:len(headers)]
                                            while len(adjusted_row) < len(headers):
                                                adjusted_row.append("")
                                            new_ws.append(adjusted_row)
                                            existing.add(adjusted_row[0])

                                    wb.remove(ws)
                                    new_ws.title = "Геодезия"
                                    ws = new_ws
                                    excel_created_or_updated = True
                                else:
                                    for row in ws.iter_rows(min_row=2, values_only=True):
                                        if row and row[0]:
                                            existing.add(row[0])
                            else:
                                raise Exception("Не удалось получить активный лист")
                        else:
                            raise Exception("Не удалось загрузить рабочую книгу")
                    except Exception as e:
                        self.log_message(f"Ошибка загрузки Excel, создаю новый: {e}")
                        wb = Workbook() if Workbook else None
                        if wb is not None:
                            ws = wb.active
                            if ws is not None:
                                ws.title = "Геодезия"
                                ws.append(headers)
                                self.adjust_columns(ws)
                                excel_created_or_updated = True
                        else:
                            self.log_message("Не удалось создать новую рабочую книгу")
                else:
                    wb = Workbook() if Workbook else None
                    if wb is not None:
                        ws = wb.active
                        if ws is not None:
                            ws.title = "Геодезия"
                            ws.append(headers)
                            self.adjust_columns(ws)
                            excel_created_or_updated = True
                    else:
                        self.log_message("Не удалось создать новую рабочую книгу")
        else:
            self.log_message("Режим: Excel отключён.")

        total_files = len(selected_filenames)
        for index, fname in enumerate(selected_filenames, 1):
            if not self.ignore_excel and EXCEL_SUPPORTED and fname in existing:
                self.log_message(f"Пропуск (уже в Excel): {fname}")
                if target_move_folder and target_move_folder != folder_path:
                    try:
                        shutil.move(os.path.join(folder_path, fname), os.path.join(target_move_folder, fname))
                        moved += 1
                    except Exception as e:
                        self.log_message(f"Не переместил {fname}: {e}")
                continue

            fpath = os.path.join(folder_path, fname)
            self.log_message(f"— Обработка: {fname} ({index}/{total_files})")
            text = self.process_pdf(fpath)
            if not text:
                if not self.ignore_excel and EXCEL_SUPPORTED and ws is not None:
                    ws.append([fname, "", "", "", "", "", "Ошибка обработки", ""])
                    excel_created_or_updated = True
                self.problem_files.append(fname)
                continue

            if self.debug_mode:
                try:
                    dbg = os.path.join(self.points_folder or folder_path, f"DEBUG_FULL_OCR_{self.sanitize_filename(fname)}.txt")
                    with open(dbg, 'w', encoding='utf-8') as f:
                        f.write(text)
                    self.log_message(f"DEBUG OCR: {dbg}")
                except Exception as e:
                    self.log_message(f"DEBUG save error: {e}")

            data = self.extract_data(text)
            found = sum(1 for v in data.values() if v)
            for k in ["Тип коммуникации", "Номер договора", "КГС", "Дата съемки"]:
                if data.get(k):
                    self.field_stats[k] += 1
            status = "Успешно" if found == 4 else ("Частично" if found > 0 else "Не распознано")

            points_status = "Нет точек"
            points_count_str = "0/0"
            if self.import_points and data.get("КГС"):
                out_folder = self.points_folder or folder_path
                if self.sort_points_by_comm:
                    out_folder = self._comm_subfolder(out_folder, data.get("Тип коммуникации"))
                points_status, points_count_str, _, _ = self.extract_and_save_coordinate_table(
                    text, data["КГС"], out_folder, fname
                )

            if not self.ignore_excel and EXCEL_SUPPORTED and ws is not None:
                ws.append([
                    fname,
                    data.get("Тип коммуникации", ""),
                    data.get("Номер договора", ""),
                    data.get("КГС", ""),
                    data.get("Дата съемки", ""),
                    points_count_str,
                    status,
                    points_status
                ])
                excel_created_or_updated = True

            self.log_message(f"Готово: {status}; точки {points_count_str}; {points_status} ({index}/{total_files})")
            processed += 1

            if target_move_folder and target_move_folder != folder_path:
                try:
                    shutil.move(fpath, os.path.join(target_move_folder, fname))
                    moved += 1
                except Exception as e:
                    self.log_message(f"Не переместил {fname}: {e}")

        excel_saved = False
        if not self.ignore_excel and EXCEL_SUPPORTED and excel_created_or_updated and wb is not None:
            if ws is not None:
                self.adjust_columns(ws)
                if ws.max_row > 1:
                    try:
                        self.apply_standard_excel_style(ws, headers)
                    except Exception as e:
                        self.log_message(f"Не удалось применить стиль Excel: {e}")
            try:
                wb.save(output_path)
                self.log_message(f"Excel сохранен: {output_path}")
                excel_saved = True
            except Exception as e:
                self.log_message(f"Excel save error: {e}")

        if self.problem_files:
            prob = os.path.join(folder_path, "проблемные_файлы.txt")
            try:
                with open(prob, 'w', encoding='utf-8') as f:
                    f.write("Проблемные файлы:\n")
                    for p in self.problem_files:
                        f.write(f"- {p}\n")
                self.log_message(f"Список проблем: {prob}")
            except Exception as e:
                self.log_message(f"Не сохранил проблемные: {e}")

        self.analyze_results(processed)
        if moved:
            self.log_message(f"Перемещено: {moved}")

        return output_path if (not self.ignore_excel and EXCEL_SUPPORTED and excel_saved) else None


# ======================= КОМПАКТНЫЙ GUI (ttk) =======================

class App(ttk.Frame):
    def __init__(self, master):
        super().__init__(master, padding=8)
        self.master = master
        self.master.title("КГС: обработчик PDF")
        self.master.geometry("980x680")
        self.master.minsize(900, 600)

        style = ttk.Style()
        try:
            style.theme_use("vista")
        except Exception:
            pass

        self.folder_path = tk.StringVar(value="")
        self.points_folder = tk.StringVar(value="")
        self.move_folder_path = tk.StringVar(value="")
        self.status_text = tk.StringVar(value="")
        self.var_sort_points = tk.BooleanVar(value=True)

        self._build_toolbar()
        self._build_body()
        self._build_bottom()

        # процессор
        self.processor = PDFProcessor(log_callback=self._append_log)

    # ---------- UI blocks ----------

    def _build_toolbar(self):
        bar = ttk.Frame(self.master)
        bar.pack(fill="x", pady=(0, 6))

        ttk.Label(bar, text="Папка:").pack(side="left")
        self.entry_folder = ttk.Entry(bar, textvariable=self.folder_path)
        self.entry_folder.pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(bar, text="Обзор…", command=self.browse_folder).pack(side="left")
        ttk.Button(bar, text="Загрузить список…", command=self.load_files_from_txt).pack(side="left", padx=(6, 0))

    def _build_body(self):
        body = ttk.Panedwindow(self.master, orient="horizontal")
        body.pack(fill="both", expand=True)

        # левая панель
        left = ttk.Frame(body)
        body.add(left, weight=3)

        head = ttk.Frame(left)
        head.pack(fill="x")
        ttk.Label(head, text="Файлы PDF").pack(side="left")
        ttk.Button(head, text="Все", width=6, command=self.select_all_files).pack(side="right")
        ttk.Button(head, text="Снять", width=6, command=self.deselect_all_files).pack(side="right", padx=(0, 6))

        # EXTENDED — стандартное поведение с Shift/Ctrl для диапазонов/множественного выбора
        self.listbox = tk.Listbox(left, selectmode=tk.EXTENDED)
        yscroll = ttk.Scrollbar(left, orient="vertical", command=self.listbox.yview)
        self.listbox.config(yscrollcommand=yscroll.set)
        self.listbox.pack(side="left", fill="both", expand=True, pady=(4, 0))
        yscroll.pack(side="right", fill="y", pady=(4, 0))

        # правая панель
        right = ttk.Frame(body, padding=(8, 0, 0, 0))
        body.add(right, weight=2)

        opt = ttk.LabelFrame(right, text="Опции")
        opt.pack(fill="x", pady=(0, 8))

        self.var_import = tk.BooleanVar(value=False)
        self.var_debug = tk.BooleanVar(value=False)
        self.var_ignore_excel = tk.BooleanVar(value=False)
        self.var_move = tk.BooleanVar(value=False)

        ttk.Checkbutton(opt, text="Импортировать координаты (TXT)", variable=self.var_import, command=self._toggle_points).pack(anchor="w")
        ttk.Checkbutton(opt, text="Разложить каталоги по типам", variable=self.var_sort_points).pack(anchor="w", padx=(18, 0))

        row_points = ttk.Frame(opt)
        row_points.pack(fill="x", pady=2)
        ttk.Label(row_points, text="Каталоги:").pack(side="left")
        self.entry_points = ttk.Entry(row_points, textvariable=self.points_folder, state="disabled")
        self.entry_points.pack(side="left", fill="x", expand=True, padx=6)
        self.btn_points = ttk.Button(row_points, text="…", width=3, command=self.browse_points, state="disabled")
        self.btn_points.pack(side="left")

        ttk.Checkbutton(opt, text="Режим отладки (сохранять OCR)", variable=self.var_debug).pack(anchor="w", pady=(4, 0))
        ttk.Checkbutton(opt, text="Игнорировать запись в Excel", variable=self.var_ignore_excel).pack(anchor="w")

        # Кнопка "Типы коммуникаций"
        ttk.Button(opt, text="Типы коммуникаций…", command=self.open_comm_types_dialog).pack(anchor="w", pady=(6, 0))

        mv = ttk.LabelFrame(right, text="Перемещение")
        mv.pack(fill="x")
        ttk.Checkbutton(mv, text="Переместить обработанные файлы", variable=self.var_move, command=self._toggle_move).pack(anchor="w")
        row_move = ttk.Frame(mv)
        row_move.pack(fill="x", pady=2)
        ttk.Label(row_move, text="Папка:").pack(side="left")
        self.entry_move = ttk.Entry(row_move, textvariable=self.move_folder_path, state="disabled")
        self.entry_move.pack(side="left", fill="x", expand=True, padx=6)
        self.btn_move = ttk.Button(row_move, text="…", width=3, command=self.browse_move, state="disabled")
        self.btn_move.pack(side="left")

    def _build_bottom(self):
        bottom = ttk.Frame(self.master)
        bottom.pack(fill="both", expand=False, pady=(6, 0))

        runrow = ttk.Frame(bottom)
        runrow.pack(fill="x")
        self.btn_run = ttk.Button(runrow, text="Запустить обработку", command=self.run_processing)
        self.btn_run.pack(side="left")
        self.btn_open_excel = ttk.Button(runrow, text="Открыть Excel", state="disabled", command=self.open_excel)
        self.btn_open_excel.pack(side="left", padx=6)
        ttk.Label(runrow, textvariable=self.status_text).pack(side="right")

        logf = ttk.LabelFrame(bottom, text="Лог")
        logf.pack(fill="both", expand=True, pady=(6, 0))
        self.log = ScrolledText(logf, wrap="word", height=10, state="disabled")
        self.log.pack(fill="both", expand=True)

    # ---------- helpers ----------

    def _append_log(self, msg):
        """
        Потокобезопасное добавление строки в лог и обновление статуса.
        Если вызвано из фонового потока — перебрасываем в главный через after.
        """
        if threading.current_thread() is not threading.main_thread():
            self.master.after(0, self._append_log, msg)
            return

        self.log.config(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.config(state="disabled")
        self.status_text.set(msg if len(msg) < 80 else msg[:77] + "…")

    def _toggle_points(self):
        if self.var_import.get():
            self.entry_points.config(state="normal")
            self.btn_points.config(state="normal")
            if not self.points_folder.get() and self.folder_path.get():
                self.points_folder.set(self.folder_path.get())
        else:
            self.entry_points.config(state="disabled")
            self.btn_points.config(state="disabled")

    def _toggle_move(self):
        if self.var_move.get():
            self.entry_move.config(state="normal")
            self.btn_move.config(state="normal")
        else:
            self.entry_move.config(state="disabled")
            self.btn_move.config(state="disabled")

    # ---------- actions ----------

    def browse_folder(self):
        path = filedialog.askdirectory()
        if not path:
            return
        self.folder_path.set(path)
        self._load_pdfs(path)
        if not self.points_folder.get():
            self.points_folder.set(path)

    def browse_points(self):
        path = filedialog.askdirectory()
        if path:
            self.points_folder.set(path)

    def browse_move(self):
        path = filedialog.askdirectory()
        if path:
            if path == self.folder_path.get():
                messagebox.showwarning("Перемещение", "Папка перемещения не может совпадать с исходной.")
                return
            self.move_folder_path.set(path)

    def _load_pdfs(self, folder):
        self.listbox.delete(0, "end")
        if not os.path.isdir(folder):
            self._append_log(f"Папка не найдена: {folder}")
            return
        files = sorted([f for f in os.listdir(folder) if f.lower().endswith(".pdf")])
        for f in files:
            self.listbox.insert("end", f)
        self._append_log(f"Найдено PDF: {len(files)}")

    def load_files_from_txt(self):
        if not self.folder_path.get():
            messagebox.showwarning("Загрузка списка", "Сначала выберите папку с PDF.")
            return
        fp = filedialog.askopenfilename(title="TXT со списком", filetypes=[("TXT", "*.txt"), ("Все файлы", "*.*")])
        if not fp:
            return
        try:
            with open(fp, 'r', encoding='utf-8') as f:
                names = {line.strip() for line in f if line.strip()}
            self.deselect_all_files()
            found = 0
            for i in range(self.listbox.size()):
                if self.listbox.get(i) in names:
                    self.listbox.selection_set(i)
                    found += 1
            miss = len(names) - found
            self._append_log(f"Из списка выбрано: {found}. Не найдено: {miss}.")
        except Exception as e:
            messagebox.showerror("Загрузка списка", f"Ошибка: {e}")

    def select_all_files(self):
        self.listbox.selection_set(0, "end")

    def deselect_all_files(self):
        self.listbox.selection_clear(0, "end")

    # ----- Диалог управления типами коммуникаций -----

    def open_comm_types_dialog(self):
        """
        Окно для настройки ожидаемых типов коммуникаций.
        Логика:
          - список всех типов,
          - выделенные строки = enabled,
          - можно добавить новый, удалить выбранные, сбросить к дефолту.
        """
        dlg = tk.Toplevel(self.master)
        dlg.title("Типы коммуникаций")
        dlg.transient(self.master)
        dlg.grab_set()
        dlg.geometry("480x400")

        main = ttk.Frame(dlg, padding=8)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text="Ожидаемые типы коммуникаций\n(выделенные строки считаются активными)").pack(anchor="w")

        list_frame = ttk.Frame(main)
        list_frame.pack(fill="both", expand=True, pady=(4, 4))

        listbox = tk.Listbox(list_frame, selectmode=tk.EXTENDED)
        listbox.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
        scrollbar.pack(side="right", fill="y")
        listbox.config(yscrollcommand=scrollbar.set)

        # заполняем из процессора
        types = self.processor.get_comm_types()
        for idx, t in enumerate(types):
            listbox.insert("end", t["name"])
            if t.get("enabled", True):
                listbox.selection_set(idx)

        # блок добавления
        add_frame = ttk.Frame(main)
        add_frame.pack(fill="x", pady=(4, 2))
        ttk.Label(add_frame, text="Новый тип:").pack(side="left")
        entry_new = ttk.Entry(add_frame)
        entry_new.pack(side="left", fill="x", expand=True, padx=4)
        def on_add():
            name = entry_new.get().strip()
            if not name:
                return
            # проверяем на дубликаты
            existing = [listbox.get(i) for i in range(listbox.size())]
            if name in existing:
                messagebox.showinfo("Типы", "Такой тип уже есть в списке.")
                return
            listbox.insert("end", name)
            # по умолчанию включаем
            listbox.selection_set(listbox.size() - 1)
            entry_new.delete(0, "end")
        ttk.Button(add_frame, text="Добавить", command=on_add).pack(side="left")

        # нижние кнопки
        btns = ttk.Frame(main)
        btns.pack(fill="x", pady=(6, 0))

        def on_delete():
            sel = list(listbox.curselection())
            if not sel:
                return
            for idx in reversed(sel):
                listbox.delete(idx)

        def on_select_all():
            listbox.selection_set(0, "end")

        def on_clear_selection():
            listbox.selection_clear(0, "end")

        def on_reset_defaults():
            # сбрасываем к дефолту в процессоре и перерисовываем список
            self.processor.reset_comm_types_to_defaults(save=True)
            listbox.delete(0, "end")
            types_local = self.processor.get_comm_types()
            for idx, t in enumerate(types_local):
                listbox.insert("end", t["name"])
                if t.get("enabled", True):
                    listbox.selection_set(idx)

        def on_save():
            all_names = [listbox.get(i) for i in range(listbox.size())]
            selected = set(listbox.curselection())
            types_new = []
            for idx, name in enumerate(all_names):
                types_new.append({
                    "name": name,
                    "enabled": (idx in selected)
                })
            self.processor.update_comm_types(types_new)
            self._append_log("Обновлён список типов коммуникаций.")
            dlg.destroy()

        ttk.Button(btns, text="Удалить выбранные", command=on_delete).pack(side="left")
        ttk.Button(btns, text="Выбрать все", command=on_select_all).pack(side="left", padx=(4, 0))
        ttk.Button(btns, text="Снять выделение", command=on_clear_selection).pack(side="left", padx=(4, 0))
        ttk.Button(btns, text="Сброс по умолчанию", command=on_reset_defaults).pack(side="left", padx=(10, 0))
        ttk.Button(btns, text="Сохранить и закрыть", command=on_save).pack(side="right")

    # ----- Запуск обработки -----

    def run_processing(self):
        """
        Теперь проверки и подготовка GUI выполняются в главном потоке,
        а тяжелая обработка уходит в фоновый поток.
        """
        folder = self.folder_path.get().strip()
        if not folder:
            messagebox.showwarning("Старт", "Выберите папку.")
            return
        if not os.path.isdir(folder):
            messagebox.showerror("Старт", "Папка не существует.")
            return

        sel_indices = self.listbox.curselection()
        if not sel_indices:
            messagebox.showwarning("Старт", "Выберите файлы.")
            return
        selected_files = [self.listbox.get(i) for i in sel_indices]

        # очистка лога и подготовка UI — в главном потоке
        self.log.config(state="normal")
        self.log.delete("1.0", "end")
        self.log.config(state="disabled")
        self.btn_open_excel.config(state="disabled")
        self.status_text.set("Работаю…")

        # конфиг процессора
        self.processor.problem_files = []
        self.processor.field_stats = defaultdict(int)
        self.processor.import_points = self.var_import.get()
        self.processor.sort_points_by_comm = self.var_sort_points.get()
        self.processor.debug_mode = self.var_debug.get()
        self.processor.ignore_excel = self.var_ignore_excel.get()
        self.processor.points_folder = self.points_folder.get().strip()

        move_folder = self.move_folder_path.get().strip() if self.var_move.get() else None

        # фоновый поток только для обработки данных
        thread = threading.Thread(
            target=self._process_files_worker,
            args=(folder, selected_files, move_folder),
            daemon=True
        )
        thread.start()

    def _process_files_worker(self, folder, selected_files, move_folder):
        try:
            result_path = self.processor.process_selected_files(folder, selected_files, move_folder)
            # После окончания обработки — все действия с GUI через after
            def on_done():
                if result_path and os.path.exists(result_path):
                    self.btn_open_excel.config(state="normal")
                    self._append_log(f"✓ Excel файл успешно создан: {os.path.basename(result_path)}")
                elif self.processor.ignore_excel:
                    self._append_log("ℹ Excel отключён — запись пропущена.")
                else:
                    self._append_log("⚠ Excel файл не был создан. Проверьте логи для диагностики.")

                self._load_pdfs(folder)
                self.status_text.set("Готово")

            self.master.after(0, on_done)

        except Exception as e:
            def on_error():
                self._append_log(f"❌ Ошибка при обработке: {e}")
                messagebox.showerror("Ошибка", f"Произошла ошибка: {e}")
                self._load_pdfs(folder)
                self.status_text.set("Готово")

            self.master.after(0, on_error)

    def open_excel(self):
        path = self.processor.output_excel_path
        if path and os.path.exists(path):
            try:
                os.startfile(path)
            except Exception as e:
                messagebox.showerror("Excel", f"Не удалось открыть: {e}")
        else:
            messagebox.showerror("Excel", "Файл не найден.")


# ======================= RUN =======================

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    app.pack(fill="both", expand=True)
    root.mainloop()
