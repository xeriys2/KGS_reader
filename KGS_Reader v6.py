from collections import defaultdict
import datetime
import json
import os
import re
import shutil
import sys
import threading
import tkinter as tk
import webbrowser
from tkinter import filedialog, messagebox
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
from datetime import datetime as dt
import locale

# Попытка установить локаль для правильной сортировки (для дат на русском)
try:
    locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'Russian_Russia.1251')
    except:
        pass

# === ИМПОРТЫ БИБЛИОТЕК ===
try:
    import fitz  # PyMuPDF
    PDF_SUPPORTED = True
except ImportError:
    fitz = None
    PDF_SUPPORTED = False
    print("Внимание: Библиотека PyMuPDF (fitz) не установлена. Обработка PDF будет недоступна.")

try:
    import pytesseract
    from PIL import Image, ImageEnhance, ImageFilter
    OCR_SUPPORTED = True
except ImportError:
    pytesseract = None
    Image = None
    ImageEnhance = None
    ImageFilter = None
    OCR_SUPPORTED = False
    print("Внимание: Библиотеки pytesseract и/или PIL не установлены. OCR будет недоступен.")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    EXCEL_SUPPORTED = True
except ImportError:
    Workbook = None
    load_workbook = None
    EXCEL_SUPPORTED = False
    print("Внимание: Библиотека openpyxl не установлена. Работа с Excel будет недоступна.")


# Ссылка на профиль в Telegram
TELEGRAM_URL = "https://t.me/Pronin_m"

# Ссылка на репозиторий (исходный код)
REPO_URL = "https://github.com/xeriys2/KGS_reader"

class ProcessingCancelled(Exception):
    """Raised when user cancels processing."""


class Tooltip:
    def __init__(self, widget, text, delay=600):
        self.widget = widget
        self.text = text
        self.delay = delay
        self._after_id = None
        self.tipwindow = None
        widget.bind("<Enter>", self._schedule)
        widget.bind("<Leave>", self._hide)
        widget.bind("<ButtonPress>", self._hide)

    def _schedule(self, _event=None):
        if self._after_id is None:
            self._after_id = self.widget.after(self.delay, self._show)

    def _show(self):
        self._after_id = None
        if self.tipwindow or not self.text:
            return
        x = self.widget.winfo_rootx() + 10
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw,
            text=self.text,
            background="#ffffe0",
            relief="solid",
            borderwidth=1,
            font=("Segoe UI", 9),
        )
        label.pack(ipadx=6, ipady=2)

    def _hide(self, _event=None):
        if self._after_id is not None:
            try:
                self.widget.after_cancel(self._after_id)
            except Exception:
                pass
            self._after_id = None
        if self.tipwindow is not None:
            try:
                self.tipwindow.destroy()
            except Exception:
                pass
            self.tipwindow = None


# ======================= БИЗНЕС-ЛОГИКА =======================
# --- ВСТАВЬ СЮДА ТВОЙ ИСХОДНЫЙ КОД PDFProcessor БЕЗ ИЗМЕНЕНИЙ ---
# (полностью как в первом файле, начиная с `class PDFProcessor:` и до конца его определения)

class PDFProcessor:
    def __init__(self, log_callback=None, progress_callback=None, cancel_event=None):
        self.log_callback = log_callback or print
        self.progress_callback = progress_callback
        self.cancel_event = cancel_event
        self.cancelled = False
        self.log_file_path = "application_log.txt"
        self.output_excel_path = ""
        self.problem_files = []
        self.field_stats = defaultdict(int)
        self.import_points = False
        self.points_folder = ""
        self.debug_mode = False
        self.ignore_excel = False
        self.sort_points_by_comm = False  # раскладывать каталоги по типам

        # Конфиг типов коммуникаций
        self.comm_types_config_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "comm_types.json"
        )
        self.default_comm_types = self._build_default_comm_types()
        self.comm_types = []  # список словарей: {"name": str, "enabled": bool}
        self.load_comm_types()

        self.setup_tesseract()

    def _build_default_comm_types(self):
        types = [
            "Кабель связи", "Тел канализация", "Кабельная канализация", "ВОЛС", "КТВ",
            "Эл кабель", "Кабель техн. и очаг. заземл", "Контур заземл", "Кабель но",
            "Водосток", "Вод-д", "Трубопровод", "Канализация хоз-быт", "ЛОС", "Дренаж",
            "Воздухопровод", "Вент. ветки", "Теплотрасса",
            "Коллектор",
            "Газопровод", "Нефтепровод", "Продуктопровод",
            "СОУЭ", "СКУД",
            "Канализация", "Нап канализация", "Сам канализация", "Водовыпуск",
            "Кабель защ", "Газ",
        ]
        seen = set()
        unique = []
        for t in types:
            if t not in seen:
                seen.add(t)
                unique.append(t)
        return unique

    def load_comm_types(self):
        if os.path.exists(self.comm_types_config_path):
            try:
                with open(self.comm_types_config_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                types = data.get("types", []) if isinstance(data, dict) else []
                cleaned = []
                names_seen = set()
                for t in types:
                    if not isinstance(t, dict):
                        continue
                    name = str(t.get("name", "")).strip()
                    enabled = bool(t.get("enabled", True))
                    if not name or name in names_seen:
                        continue
                    names_seen.add(name)
                    cleaned.append({"name": name, "enabled": enabled})
                if cleaned:
                    self.comm_types = cleaned
                else:
                    self.reset_comm_types_to_defaults(save=False)
            except Exception as e:
                self.log_message(f"Не удалось загрузить comm_types.json: {e}. Использую значения по умолчанию.")
                self.reset_comm_types_to_defaults(save=False)
        else:
            self.reset_comm_types_to_defaults(save=False)

    def save_comm_types(self):
        try:
            data = {"types": self.comm_types}
            with open(self.comm_types_config_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self.log_message(f"Типы коммуникаций сохранены: {self.comm_types_config_path}")
        except Exception as e:
            self.log_message(f"Ошибка сохранения comm_types.json: {e}")

    def reset_comm_types_to_defaults(self, save=True):
        self.comm_types = [{"name": name, "enabled": True} for name in self.default_comm_types]
        if save:
            self.save_comm_types()

    def get_comm_types(self):
        return [dict(name=t["name"], enabled=bool(t.get("enabled", True))) for t in self.comm_types]

    def update_comm_types(self, types_list):
        cleaned = []
        names_seen = set()
        for t in types_list:
            if not isinstance(t, dict):
                continue
            name = str(t.get("name", "")).strip()
            if not name or name in names_seen:
                continue
            enabled = bool(t.get("enabled", True))
            names_seen.add(name)
            cleaned.append({"name": name, "enabled": enabled})
        self.comm_types = cleaned
        self.save_comm_types()

    def get_allowed_comm_types(self):
        if not self.comm_types:
            return self.default_comm_types
        return [t["name"] for t in self.comm_types if t.get("enabled", True)]

    def set_progress_callback(self, cb):
        self.progress_callback = cb

    def set_cancel_event(self, event):
        self.cancel_event = event

    def _report_progress(self, **info):
        if not self.progress_callback:
            return
        try:
            self.progress_callback(**info)
        except Exception:
            pass

    def check_cancelled(self):
        if self.cancel_event is not None and self.cancel_event.is_set():
            self.cancelled = True
            raise ProcessingCancelled()

    def log_message(self, message):
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{ts}] {message}\n"
        self.log_callback(message)
        try:
            if self.log_file_path:
                with open(self.log_file_path, "a", encoding="utf-8") as f:
                    f.write(line)
        except Exception as e:
            self.log_callback(f"ОШИБКА логирования: {e}")
    def setup_tesseract(self):
        if not OCR_SUPPORTED or pytesseract is None:
            self.log_message("OCR не поддерживается: библиотеки не установлены")
            return
        try:
            if sys.platform == 'win32':
                paths = [
                    r'C:\Program Files\Tesseract-OCR\tesseract.exe',
                    r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
                    os.path.join(os.environ.get('LOCALAPPDATA', ''), 'Tesseract-OCR', 'tesseract.exe'),
                    os.path.join(os.path.expanduser('~'), 'AppData', 'Local', 'Tesseract-OCR', 'tesseract.exe')
                ]
                for p in paths:
                    if os.path.exists(p):
                        pytesseract.pytesseract.tesseract_cmd = p
                        break
            pytesseract.get_tesseract_version()
            self.log_message("Tesseract OCR: ок")
        except Exception as e:
            messagebox.showerror("Tesseract", f"Tesseract не найден/не работает: {e}")
            self.log_message(f"Tesseract ошибка: {e}")
    def enhance_image(self, image):
        if not OCR_SUPPORTED or ImageEnhance is None or ImageFilter is None:
            return image
        image = ImageEnhance.Contrast(image).enhance(2.0)
        image = ImageEnhance.Sharpness(image).enhance(2.0)
        image = image.convert('L')
        image = image.filter(ImageFilter.MedianFilter(size=3))
        image = image.point(lambda x: 0 if x < 140 else 255)
        return image
    def extract_text_with_ocr(self, page):
        self.check_cancelled()
        if not OCR_SUPPORTED or not PDF_SUPPORTED or fitz is None or pytesseract is None or Image is None:
            self.log_message("OCR не поддерживается: необходимые библиотеки не установлены")
            return ""
        try:
            matrix = fitz.Matrix(4, 4)
            pix = page.get_pixmap(matrix=matrix, dpi=300)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img = self.enhance_image(img)
            return pytesseract.image_to_string(img, config=r'--oem 3 --psm 3 -l rus+eng --dpi 300')
        except Exception as e:
            self.log_message(f"OCR ошибка: {e}")
            return ""
    def process_pdf(self, file_path):
        if not PDF_SUPPORTED or fitz is None:
            self.log_message("Обработка PDF не поддерживается: библиотека PyMuPDF не установлена")
            return None
        doc = None
        try:
            doc = fitz.open(file_path)
            total_pages = getattr(doc, "page_count", None) or len(doc)
            full_text = ""
            base = os.path.basename(file_path)
            for i, page in enumerate(doc):
                self.check_cancelled()
                self._report_progress(
                    filename=base,
                    page_index=i + 1,
                    total_pages=total_pages,
                )
                text = page.get_text("text")
                if not text or len(text.strip()) < 50:
                    self.log_message(f"Стр.{i+1}: OCR")
                    text = self.extract_text_with_ocr(page)
                full_text += text + "\n"
            return full_text
        except ProcessingCancelled:
            raise
        except Exception as e:
            self.log_message(f"Ошибка PDF {os.path.basename(file_path)}: {e}")
            return None
        finally:
            if doc is not None:
                try:
                    doc.close()
                except Exception:
                    pass
    def extract_kgc_number(self, text):
        patterns = [
            r"№ КГС:\s*(\d{2,5}[-\/]\d{2,5})",
            r"№ КГС:\s*([A-ZА-Я]?\d+[A-ZА-Я]?)",
            r"(?:КГС|№|N)\s*[:\-]?\s*(\d{2,5}[-\/]\d{2,5})",
            r"(?:КГС|№|N)\s*[:\-]?\s*([A-ZА-Я]?\d+[A-ZА-Я]?)",
            r"КГС\s*([^\n,;]+)",
            r"\b(\d{2,5}-\d{2,5})\b",
            r"\b(\d{5}-\d{2})\b"
        ]
        for p in patterns:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                kgs = (m.group(1) if m.groups() else m.group(0)).strip()
                kgs = re.sub(r'[^\dА-ЯA-Z\-/]', '', kgs)
                if len(kgs) >= 4:
                    return kgs
        return None
    def similarity(self, a, b):
        a, b = a.lower(), b.lower()
        if a in b or b in a:
            return 0.9
        a_words = set(a.split())
        b_words = set(b.split())
        if not a_words or not b_words:
            return 0.0
        common_words = a_words & b_words
        union_words = a_words | b_words
        similarity_score = len(common_words) / len(union_words)
        if len(a) > 3 and len(b) > 3:
            a_short = ''.join([word[0] for word in a.split() if word])
            b_short = ''.join([word[0] for word in b.split() if word])
            if a_short and b_short and a_short == b_short:
                similarity_score = max(similarity_score, 0.7)
        return similarity_score
    def normalize_communication_type(self, text, allowed_communications):
        if not text:
            return None
        text = text.strip()
        normalization_map = {
            r'кабел[ьи]?\с*связ[и]?': 'Кабель связи',
            r'тел[е]?\с*канализац[ия]{2,4}': 'Тел канализация',
            r'кабел[ьи]?\с*канализац[ия]{2,4}': 'Кабельная канализация',
            r'волоконно-?\с*оптическ[ая]?\с*лини[яи]?\с*связ[и]?': 'ВОЛС',
            r'кабел[ьное]?\с*телевидени[е]?': 'КТВ',
            r'эл[ек]?\с*кабел[ья]?': 'Эл кабель',
            r'кабел[ь]?\с*техн[\.]?\с*и\s*очаг[\.]?\с*заземл[ения]?': 'Кабель техн. и очаг. заземл',
            r'контур\s*заземл[ения]?': 'Контур заземл',
            r'кабел[ь]?\с*н[оo0]': 'Кабель но',
            r'наружн[ое]?\с*освещени[е]': 'Кабель но',
            r'ливнев[ая]?\с*канализац[ия]{2,4}': 'Водосток',
            r'вод-?д': 'Вод-д',
            r'водопровод': 'Вод-д',
            r'трубопровод': 'Трубопровод',
            r'канализац[ия]{2,4}\с*хоз-?быт': 'Канализация',
            r'хоз-?бытов[ая]?\с*канализац[ия]{2,4}': 'Канализация',
            r'лос\b': 'ЛОС',
            r'локальн[ые]?\с*очистн[ые]?\с*сооружени[я]': 'ЛОС',
            r'дренаж': 'Дренаж',
            r'воздухопровод': 'Воздухопровод',
            r'вент[\.]?\s*ветк[и]?': 'Вент. ветки',
            r'вентиляционн[ые]?\с*ветк[и]?': 'Вент. ветки',
            r'теплотрасса': 'Теплотрасса',
            r'теплов[ые]?\с*сет[и]?': 'Теплотрасса',
            r'коллектор': 'Коллектор',
            r'газопровод': 'Газопровод',
            r'нефтепровод': 'Нефтепровод',
            r'продуктопровод': 'Продуктопровод',
            r'соуэ\b': 'СОУЭ',
            r'систем[аы]?\с*оповещени[я]?\с*и\s*управлен[ие]?\с*эвакуаци[ей]': 'СОУЭ',
            r'скуд\b': 'СКУД',
            r'систем[аы]?\с*контрол[я]?\с*управлени[я]?\с*доступом': 'СКУД',
            r'\bканализац(ия)?\b': 'Канализация',
            r'нап[оо]рн[аяые]?\с*канализац[ия]{2,4}|нап\.?\s*канализац': 'Нап канализация',
            r'сам[о]?течн[аяые]?\с*канализац[ия]{2,4}|сам\.?\s*канализац': 'Сам канализация',
            r'водо[вв]ыпуск': 'Водовыпуск',
            r'кабел[ьи]?\с*защ': 'Кабель защ',
            r'\bгаз\b': 'Газ',
            r'газопровод': 'Газопровод',
            r'тепл(о|\.|\b)': 'Теплотрасса',
        }
        for pattern, standard in normalization_map.items():
            if re.search(pattern, text, re.IGNORECASE):
                if not allowed_communications or standard in allowed_communications:
                    return standard
        for comm_type in allowed_communications:
            if self.similarity(text.lower(), comm_type.lower()) > 0.9:
                return comm_type
        return None
    def find_best_communication_match(self, text, allowed_communications):
        best_match = None
        best_score = 0.6
        words_to_check = []
        lines = text.split('\n')
        for line in lines:
            words = line.strip().split()
            for i in range(len(words) - 1):
                phrase = ' '.join(words[i:i+2])
                if len(phrase) > 5:
                    words_to_check.append(phrase)
                if i < len(words) - 2:
                    phrase = ' '.join(words[i:i+3])
                    if len(phrase) > 8:
                        words_to_check.append(phrase)
        for phrase in words_to_check:
            for comm_type in allowed_communications:
                score = self.similarity(phrase.lower(), comm_type.lower())
                if score > best_score:
                    best_score = score
                    best_match = comm_type
        if best_match:
            self.log_message(f"Определен тип коммуникации: '{best_match}' (схожесть: {best_score:.1%})")
            return best_match
        return None
    def extract_field(self, text, field_name, patterns):
        if field_name == "Тип коммуникации":
            allowed_communications = self.get_allowed_comm_types()
            patterns += [
                r"Кабель\s*связи\b", r"Тел\s*канализац[ия]{2,4}\b", r"Кабельная\s*канализац[ия]{2,4}\b",
                r"ВОЛС\b", r"КТВ\b", r"Эл\s*кабель\b", r"Кабель\s*техн\.?\s*и\s*очаг\.?\s*заземл\b",
                r"Контур\s*заземл\b", r"Кабель\s*н[оo0]\b", r"Водосток\b", r"Вод-?д\b", r"Трубопровод\b",
                r"Канализац[ия]{2,4}\s*хоз-?быт\b", r"ЛОС\b", r"Дренаж\b", r"Воздухопровод\b",
                r"Вент\.?\s*ветки\b", r"Теплотрасса\b", r"Коллектор\b", r"Газопровод\b", r"Нефтепровод\b",
                r"Продуктопровод\b", r"СОУЭ\b", r"СКУД\b", r'\bКанализац(ия)?\b', r'Нап\s*канализац',
                r'Сам\s*канализац', r'Водовыпуск\b', r'Кабель\s*защ\b', r'\bГаз\b', r'\bТепл\b'
            ]
            patterns += [rf"{re.escape(ct)}\b" for ct in allowed_communications if ct]
            for p in patterns:
                m = re.search(p, text, re.IGNORECASE)
                if m:
                    found_text = (m.group(1) if (m.groups() and m.group(1)) else m.group(0)).strip()
                    normalized = self.normalize_communication_type(found_text, allowed_communications)
                    if normalized:
                        return normalized
            best_match = self.find_best_communication_match(text, allowed_communications)
            if best_match:
                return best_match
            return None
        for p in patterns:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                val = (m.group(1) if (m.groups() and m.group(1)) else m.group(0)).strip()
                if field_name == "Номер договора":
                    val = re.sub(r'[^0-9A-Za-zА-Яа-я\/\-]', '', val)
                    m2 = re.search(
                        r'(\b\d+\/[A-ZА-Я]+\/[\wА-Яа-я]+\-?\d+\/\d+\b|\b\d+\/[A-ZА-Я]+\-?\d+\/\d+\b|\b\d+\/\d+\-?\d+\b|\b\d+\/[A-ZА-Я]+\-\d+\b|\b\d{1,2}\/\d{5}\-?\d{1,2}\b|\b\d{1,2}\/\d{5}\b|\b[A-ZА-Я]+\-\d+\/\d+\b|\b\d{1,5}[A-ZА-Я]*[-\/]\d{1,5}\b)',
                        val, re.IGNORECASE
                    )
                    return m2.group(1) if m2 else val
                return val
        return None
    def extract_data(self, text):
        data = {}
        data["Тип коммуникации"] = self.extract_field(text, "Тип коммуникации", [
            r"Вид\s*коммуникации/здания,\s*сооружения:\s*([^\n]+)",
            r"Вид\s*коммуникации[^\n:]*[:\s]*([^\n]+)",
            r"Тип\s*коммуникации[^\n:]*[:\s]*([^\n]+)",
            r"Коммуникац[ияи][^\n:]*[:\s]*([^\n]+)",
        ])
        data["Номер договора"] = self.extract_field(text, "Номер договора", [
            r"№\s*договора\s*\(?соглашения\)?\s*на\s*проведение\s*работ[^\n:]*[:\s]*([^\n,;]+)",
            r"№\s*договора[^\n:]*[:\s]*([^\n,;]+)",
            r"Договор\s*№\s*(\S+)",
            r"№\s*контракта[^\n:]*[:\s]*(\S+)"
        ])
        data["КГС"] = self.extract_kgc_number(text)
        data["Дата съемки"] = self.extract_field(text, "Дата съемки", [
            r"Дата\s*съ[её]мки\s*[:\s]*([0-9]{2}\.[0-9]{2}\.[0-9]{4})",
            r"Съемка\s*от\s*([0-9]{2}\.[0-9]{2}\.[0-9]{4})",
            r"\b\d{2}\.\d{2}\.\d{4}\b"
        ])
        return data
    def sanitize_filename(self, s):
        if not s: return "UNKNOWN_KGS"
        s = re.sub(r'[\\/*?:"<>|]', '_', str(s).strip())
        return re.sub(r'\s+', '_', s)
    def _comm_subfolder(self, base_folder, comm_type):
        name = self.sanitize_filename(comm_type or "Без_типа")
        return os.path.join(base_folder, name)
    def extract_and_save_coordinate_table(self, document_text, kgs, out_folder, src_pdf):
        if not kgs:
            return "Нет КГС", "0/0", 0, 0
        fname = os.path.join(out_folder, f"{self.sanitize_filename(kgs)}.txt")
        issues_name = os.path.join(out_folder, f"{self.sanitize_filename(kgs)}_issues.txt")
        start_keys = [
            r"каталог\s+(?:исполнительных\s+|фактических\s+)?координат",
            r"ведомость\s+(?:исполнительных\s+|фактических\s+)?координат",
            r"координаты\s+точек", r"координаты\s+пунктов",
            r"№\s*точки.*X.*[YУ].*[HН]?", r"n/n\s*по\s*съемке\s*[xх]\s*,\s*[мМm]\s*[yу]\s*,\s*[мМm]\s*[hн]\s*,\s*[мМm]"
        ]
        row_re = re.compile(
            r"^\s*(\d+)(?:\s+\d+)?\s+([-–—−]?\d{1,3}(?:\s*\d{3})*[.,]\d{1,3})\s+([-–—−]?\d{1,3}(?:\s*\d{3})*[.,]\d{1,3})(?:\s+([-–—−]?\d{1,3}(?:\s*\d{3})*[.,]\d{1,3}))?(?:\s+(.*))?$",
            re.IGNORECASE
        )
        num_token_re = re.compile(r"[-–—−]?\d+(?:\s?\d{3})*(?:[.,]\d+)?")
        def clean_num_string(v: str) -> str:
            if not v: return ""
            s = str(v).strip()
            repl = {'O':'0','О':'0','I':'1','L':'1','Е':'1','Z':'2','З':'3','S':'5','Б':'6',
                    'B':'8','В':'8','°':'0','=':'-','−':'-','–':'-','—':'-','_':'-'}
            for a,b in repl.items(): s = s.replace(a,b)
            s = re.sub(r'\s+', '', s).replace(',', '.')
            if s.count('-')>1:
                s = ('-'+s.replace('-','')) if s.startswith('-') else s.replace('-','')
            if s.count('.')>1:
                p = s.split('.'); s = p[0]+'.'+''.join(p[1:])
            s = re.sub(r'[^\d\.-]', '', s)
            return s if re.search(r'\d', s) else ""
        def as_float(s):
            try:
                return float(s)
            except:
                return None
        def fuzzy_parse(line: str):
            m_id = re.match(r"\s*(\d+)(?:\s+\d+)?", line)
            if not m_id:
                return None
            pid = m_id.group(1)
            rest = line[m_id.end():]
            nums = [clean_num_string(t.group(0)) for t in num_token_re.finditer(rest)]
            nums = [n for n in nums if n]
            x = y = h = ""
            if len(nums) >= 2:
                x, y = nums[0], nums[1]
                if len(nums) >= 3:
                    h = nums[2]
            desc = ""
            if nums:
                last = None
                for m in num_token_re.finditer(rest):
                    last = m
                if last:
                    desc = rest[last.end():].strip()
            else:
                desc = rest.strip()
            return pid, x, y, h, desc
        lines = document_text.splitlines()
        raw_rows = []
        parsing = False
        skip = 0
        MAX_SKIP = 10
        issues = []
        max_id = 0
        table_header_found = False
        for line in lines:
            t = line.strip()
            if not t:
                if parsing: 
                    skip += 1
                    if skip >= MAX_SKIP and raw_rows:
                        break
                continue
            if not parsing:
                if any(re.search(k, t, re.IGNORECASE) for k in start_keys):
                    parsing, skip = True, 0
                continue
            if parsing and not table_header_found:
                table_header_patterns = [
                    r"n/n\s*по\s*съемке\s*[xх]\s*,\s*[мmМ]\s*[yу]\s*,\s*[мmМ]\s*[hн]\s*,\s*[мmМ]?",
                    r"n/n\s*по\s*съемке\s*[xх]\s*,?\s*[мmМ]?\s*[yу]\s*,?\s*[мmМ]?\s*[hн]\s*,?\s*[мmМ]?",
                    r"№\s*точки.*[xхyуhн]",
                    r"№\s*точки.*координаты"
                ]
                if any(re.search(pattern, t, re.IGNORECASE) for pattern in table_header_patterns):
                    table_header_found = True
                    skip = 0
                continue
            m = row_re.match(t)
            if m:
                skip = 0
                pid, x, y, h, d = m.groups()
                x, y, h = clean_num_string(x), clean_num_string(y), clean_num_string(h)
            else:
                skip += 1
                parsed = fuzzy_parse(t)
                if not parsed:
                    if raw_rows and skip >= MAX_SKIP:
                        break
                    else:
                        continue
                pid, x, y, h, d = parsed
                x, y, h = clean_num_string(x), clean_num_string(y), clean_num_string(h)
            if pid and pid.isdigit():
                max_id = max(max_id, int(pid))
            raw_rows.append({"pid": pid, "x": x, "y": y, "h": h, "d": (d or "").strip(), "line": t})
        if not raw_rows:
            self.log_message(f"Каталог координат не найден ({src_pdf})")
            return "Нет точек", "0/0", 0, 0
        xs = [as_float(r["x"]) for r in raw_rows if r["x"]]
        ys = [as_float(r["y"]) for r in raw_rows if r["y"]]
        hs = [as_float(r["h"]) for r in raw_rows if r["h"]]
        def majority_negative(values):
            vs = [v for v in values if v is not None]
            if not vs: return False
            neg = sum(1 for v in vs if v < 0)
            pos = sum(1 for v in vs if v > 0)
            return neg > pos
        x_should_be_negative = majority_negative(xs)
        y_should_be_negative = majority_negative(ys)
        fixed_rows = []
        for r in raw_rows:
            pid, x_s, y_s, h_s, d = r["pid"], r["x"], r["y"], r["h"], r["d"]
            x, y, h = as_float(x_s), as_float(y_s), as_float(h_s)
            notes = []
            if x is None or y is None:
                notes.append("неполная строка (нет X или Y)")
            def maybe_fix_sign(val, s, should_neg, label):
                if val is None: 
                    return val, s, False
                if should_neg and val > 0 and abs(val) > 500:
                    s_fixed = "-" + s if not s.startswith("-") else s
                    notes.append(f"возможен потерянный минус у {label} -> исправил")
                    return -val, s_fixed, True
                return val, s, False
            x, x_s, _ = maybe_fix_sign(x, x_s, x_should_be_negative, "X")
            y, y_s, _ = maybe_fix_sign(y, y_s, y_should_be_negative, "Y")
            if h is not None and h > 500:
                if re.fullmatch(r"\d{4,6}", h_s or ""):
                    h_s_fixed = h_s[:-2] + "." + h_s[-2:]
                    h_fixed = as_float(h_s_fixed)
                    if h_fixed is not None and 0 < h_fixed < 500:
                        notes.append(f"высота без точки ({h_s}) -> {h_s_fixed}")
                        h, h_s = h_fixed, h_s_fixed
            if ("неполная строка" in " ".join(notes)) or x is None or y is None:
                issues.append(f"{pid}\t{x_s or ''}\t{y_s or ''}\t{h_s or ''}\t{d}    <-- ПРОБЛЕМА: {', '.join(notes) or 'не удалось распарсить'}")
            fixed_rows.append((pid, x_s or "", y_s or "", h_s or "", d))
        os.makedirs(out_folder, exist_ok=True)
        with open(fname, 'w', encoding='utf-8') as f:
            for pid, x_s, y_s, h_s, d in fixed_rows:
                if x_s and y_s:
                    f.write(f"{pid}\t{x_s}\t{y_s}\t{h_s}\t{d}\n")
        if issues:
            with open(issues_name, 'w', encoding='utf-8') as f:
                f.write("Строки с подозрениями/ошибками (после грубых автоправок):\n")
                for it in issues:
                    f.write(it + "\n")
            self.log_message(f"⚠ Обнаружены проблемные строки: {os.path.basename(issues_name)}")
        cnt = sum(1 for pid, x_s, y_s, h_s, d in fixed_rows if x_s and y_s)
        count_str = f"{cnt}/{max_id if max_id else cnt}"
        self.log_message(f"Каталог: {os.path.basename(fname)} | Точек: {count_str}")
        return "Точки сохранены", count_str, cnt, max_id
    def adjust_columns(self, ws):
        if not EXCEL_SUPPORTED:
            return
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            width = min(max((len(str(c.value)) for c in col if c.value), default=0) + 2, 50)
            ws.column_dimensions[get_column_letter(col[0].column)].width = width
    def apply_standard_excel_style(self, ws, headers):
        if not EXCEL_SUPPORTED:
            return
        try:
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            body_font = Font(size=11, name="Calibri")
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(style="thin", color="9E9E9E")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            max_row = ws.max_row
            max_col = ws.max_column
            if max_row < 1 or max_col < 1:
                return
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = border
            ws.row_dimensions[1].height = 22
            for r in range(2, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = body_font
                    cell.alignment = center
                    cell.border = border
            try:
                date_col_idx = headers.index("Дата съемки") + 1
                for r in range(2, max_row + 1):
                    ws.cell(row=r, column=date_col_idx).number_format = "DD.MM.YYYY"
            except ValueError:
                pass
            ws.freeze_panes = "A2"
            if max_row > 1:
                ws.auto_filter.ref = ws.dimensions
        except Exception as e:
            self.log_message(f"Не удалось применить стиль Excel: {e}")
    def analyze_results(self, total):
        self.log_message("--- Сводка ---")
        if not total:
            self.log_message("Нет обработанных файлов.")
            return
        for f in ["Тип коммуникации", "Номер договора", "КГС", "Дата съемки"]:
            c = self.field_stats.get(f, 0)
            self.log_message(f"{f}: {c}/{total} ({(c/total*100):.1f}%)")
    def process_selected_files(self, folder_path, selected_filenames, target_move_folder=None):
        if not os.path.exists(folder_path):
            messagebox.showerror("Ошибка", "Папка не существует!")
            return None
        output_path = os.path.join(folder_path, "Реестр_геодезических_съемок.xlsx")
        self.output_excel_path = output_path
        self.log_file_path = os.path.join(folder_path, "application_log.txt")
        headers = ["Файл", "Тип коммуникации", "Номер договора", "КГС", "Дата съемки", "Количество точек", "Статус", "Точки"]
        wb = ws = None
        existing = set()
        processed = 0
        moved = 0
        excel_created_or_updated = False
        if not self.ignore_excel:
            if not EXCEL_SUPPORTED:
                self.log_message("Работа с Excel не поддерживается: библиотека openpyxl не установлена")
            else:
                if os.path.exists(output_path):
                    try:
                        wb = load_workbook(output_path)
                        if wb is not None:
                            ws = wb.active
                            if ws is not None:
                                current_headers = [cell.value for cell in ws[1] if cell.value]
                                if current_headers != headers:
                                    self.log_message("Обновляю структуру Excel файла...")
                                    new_ws = wb.create_sheet("Геодезия_новый")
                                    new_ws.append(headers)
                                    for row in ws.iter_rows(min_row=2, values_only=True):
                                        if row and row[0]:
                                            adjusted_row = list(row)[:len(headers)]
                                            while len(adjusted_row) < len(headers):
                                                adjusted_row.append("")
                                            new_ws.append(adjusted_row)
                                            existing.add(adjusted_row[0])
                                    wb.remove(ws)
                                    new_ws.title = "Геодезия"
                                    ws = new_ws
                                    excel_created_or_updated = True
                                else:
                                    for row in ws.iter_rows(min_row=2, values_only=True):
                                        if row and row[0]:
                                            existing.add(row[0])
                            else:
                                raise Exception("Не удалось получить активный лист")
                    except Exception as e:
                        self.log_message(f"Ошибка загрузки Excel, создаю новый: {e}")
                        wb = Workbook()
                        if wb is not None:
                            ws = wb.active
                            if ws is not None:
                                ws.title = "Геодезия"
                                ws.append(headers)
                                self.adjust_columns(ws)
                                excel_created_or_updated = True
                else:
                    wb = Workbook()
                    if wb is not None:
                        ws = wb.active
                        if ws is not None:
                            ws.title = "Геодезия"
                            ws.append(headers)
                            self.adjust_columns(ws)
                            excel_created_or_updated = True
        else:
            self.log_message("Режим: Excel отключён.")
        total_files = len(selected_filenames)
        cancelled = False
        for index, fname in enumerate(selected_filenames, 1):
            try:
                self.check_cancelled()
                self._report_progress(
                    file_index=index,
                    total_files=total_files,
                    filename=fname,
                    page_index=0,
                    total_pages=0,
                )
                if not self.ignore_excel and EXCEL_SUPPORTED and fname in existing:
                    self.log_message(f"Пропуск (уже в Excel): {fname}")
                    if target_move_folder and target_move_folder != folder_path:
                        try:
                            shutil.move(os.path.join(folder_path, fname), os.path.join(target_move_folder, fname))
                            moved += 1
                        except Exception as e:
                            self.log_message(f"Не переместил {fname}: {e}")
                    continue
                fpath = os.path.join(folder_path, fname)
                self.log_message(f"— Обработка: {fname} ({index}/{total_files})")
                text = self.process_pdf(fpath)
                if not text:
                    if not self.ignore_excel and EXCEL_SUPPORTED and ws is not None:
                        ws.append([fname, "", "", "", "", "", "Ошибка обработки", ""])
                        excel_created_or_updated = True
                    self.problem_files.append(fname)
                    continue
            except ProcessingCancelled:
                cancelled = True
                self.log_message("Отмена пользователем. Останавливаю обработку.")
                break
            if self.debug_mode:
                try:
                    dbg = os.path.join(self.points_folder or folder_path, f"DEBUG_FULL_OCR_{self.sanitize_filename(fname)}.txt")
                    with open(dbg, 'w', encoding='utf-8') as f:
                        f.write(text)
                    self.log_message(f"DEBUG OCR: {dbg}")
                except Exception as e:
                    self.log_message(f"DEBUG save error: {e}")
            data = self.extract_data(text)
            found = sum(1 for v in data.values() if v)
            for k in ["Тип коммуникации", "Номер договора", "КГС", "Дата съемки"]:
                if data.get(k):
                    self.field_stats[k] += 1
            status = "Успешно" if found == 4 else ("Частично" if found > 0 else "Не распознано")
            points_status = "Нет точек"
            points_count_str = "0/0"
            if self.import_points and data.get("КГС"):
                out_folder = self.points_folder or folder_path
                if self.sort_points_by_comm:
                    out_folder = self._comm_subfolder(out_folder, data.get("Тип коммуникации"))
                points_status, points_count_str, _, _ = self.extract_and_save_coordinate_table(
                    text, data["КГС"], out_folder, fname
                )
            if not self.ignore_excel and EXCEL_SUPPORTED and ws is not None:
                ws.append([
                    fname,
                    data.get("Тип коммуникации", ""),
                    data.get("Номер договора", ""),
                    data.get("КГС", ""),
                    data.get("Дата съемки", ""),
                    points_count_str,
                    status,
                    points_status
                ])
                excel_created_or_updated = True
            self.log_message(f"Готово: {status}; точки {points_count_str}; {points_status} ({index}/{total_files})")
            processed += 1
            if target_move_folder and target_move_folder != folder_path:
                try:
                    shutil.move(fpath, os.path.join(target_move_folder, fname))
                    moved += 1
                except Exception as e:
                    self.log_message(f"Не переместил {fname}: {e}")
        excel_saved = False
        if not self.ignore_excel and EXCEL_SUPPORTED and excel_created_or_updated and wb is not None:
            if ws is not None:
                self.adjust_columns(ws)
                if ws.max_row > 1:
                    try:
                        self.apply_standard_excel_style(ws, headers)
                    except Exception as e:
                        self.log_message(f"Не удалось применить стиль Excel: {e}")
            try:
                wb.save(output_path)
                self.log_message(f"Excel сохранен: {output_path}")
                excel_saved = True
            except PermissionError:
                error_msg = f"Не удалось сохранить Excel файл: {output_path}. Файл может быть открыт в Excel. Закройте файл и попробуйте снова."
                messagebox.showerror("Ошибка сохранения", error_msg)
                self.log_message(f"Ошибка сохранения Excel: {error_msg}")
            except Exception as e:
                error_msg = f"Не удалось сохранить Excel файл: {e}"
                messagebox.showerror("Excel", error_msg)
                self.log_message(f"Excel save error: {error_msg}")
        if self.problem_files:
            prob = os.path.join(folder_path, "проблемные_файлы.txt")
            try:
                with open(prob, 'w', encoding='utf-8') as f:
                    f.write("Проблемные файлы:\n")
                    for p in self.problem_files:
                        f.write(f"- {p}\n")
                self.log_message(f"Список проблем: {prob}")
            except Exception as e:
                self.log_message(f"Не сохранил проблемные: {e}")
        self.analyze_results(processed)
        if moved:
            self.log_message(f"Перемещено: {moved}")
        if cancelled:
            self.log_message(f"Обработка остановлена пользователем: {processed}/{total_files} файлов.")
        return output_path if (not self.ignore_excel and EXCEL_SUPPORTED and excel_saved) else None


# ======================= УЛУЧШЕННЫЙ ИНТЕРФЕЙС =======================

class ModernFileSelector(ttk.Frame):
    def __init__(self, parent, on_selection_change=None):
        super().__init__(parent)
        self.on_selection_change = on_selection_change
        self.files_data = []
        self.file_index = {}
        self.sort_column = 'date'
        self.sort_reverse = True
        
        self._create_widgets()
        self._create_context_menu()
        self._setup_bindings()
    
    def _create_widgets(self):
        toolbar = ttk.Frame(self)
        toolbar.pack(fill='x', pady=(0, 5))
        
        ttk.Button(toolbar, text="ВСЕ", width=8, command=self.select_all).pack(side='left', padx=2)
        ttk.Button(toolbar, text="НИЧЕГО", width=8, command=self.select_none).pack(side='left', padx=2)
        ttk.Button(toolbar, text="ИНВЕРТ.", width=8, command=self.invert_selection).pack(side='left', padx=2)
        
        ttk.Separator(toolbar, orient='vertical').pack(side='left', padx=10, fill='y')
        
        ttk.Label(toolbar, text="Сортировать:").pack(side='left', padx=(0, 5))
        ttk.Button(toolbar, text="По имени", command=lambda: self.sort_by_column('name')).pack(side='left', padx=2)
        ttk.Button(toolbar, text="По дате ↓", command=lambda: self.sort_by_column('date')).pack(side='left', padx=2)
        ttk.Button(toolbar, text="По дате ↑", command=lambda: self.sort_by_column('date', reverse=False)).pack(side='left', padx=2)
        
        search_frame = ttk.Frame(toolbar)
        search_frame.pack(side='right')
        ttk.Label(search_frame, text="Поиск:").pack(side='left')
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=20)
        self.search_entry.pack(side='left', padx=5)
        self.search_entry.bind('<KeyRelease>', lambda e: self._apply_filter())
        
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill='both', expand=True)
        
        columns = ('selected', 'name', 'size', 'date', 'time')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', selectmode='extended')
        
        col_config = [
            ('selected', '✓', 45, 'center'),
            ('name', 'Имя файла', 350, 'w'),
            ('size', 'Размер (KB)', 80, 'center'),
            ('date', 'Дата', 100, 'center'),
            ('time', 'Время', 80, 'center')
        ]
        
        for col_id, heading, width, anchor in col_config:
            self.tree.heading(col_id, text=heading)
            self.tree.column(col_id, width=width, anchor=anchor)
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        
        self.tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        
        self.status_var = tk.StringVar(value="Файлов: 0 | Выбрано: 0")
        ttk.Label(self, textvariable=self.status_var).pack(side='bottom', fill='x')
    
    def _create_context_menu(self):
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(
            label="Отметить выбранные",
            command=lambda: self._set_items_state(self.tree.selection(), True)
        )
        self.context_menu.add_command(
            label="Снять отметки",
            command=lambda: self._set_items_state(self.tree.selection(), False)
        )
        self.context_menu.add_separator()
        self.context_menu.add_command(
            label="Инвертировать выбранные",
            command=lambda: self._invert_items_state(self.tree.selection())
        )
    
    def _setup_bindings(self):
        self.tree.bind('<Button-1>', self._on_click)
        self.tree.bind('<Button-3>', self._on_right_click)
        self.tree.bind('<Control-a>', lambda e: self.select_all())
        self.tree.bind('<Control-Shift-A>', lambda e: self.select_none())
        self.tree.bind('<Control-i>', lambda e: self.invert_selection())
    
    def _get_file_info_by_item(self, item_id):
        if item_id in self.file_index:
            return self.file_index[item_id]
        
        values = self.tree.item(item_id, 'values')
        if values and len(values) > 1:
            name = values[1]
            for info in self.files_data:
                if info['filename'] == name:
                    return info
        return None
    
    def _set_items_state(self, items, state):
        if not items:
            return
        
        checkmark = "✓" if state else "○"
        changed = False
        
        for item in items:
            info = self._get_file_info_by_item(item)
            if not info:
                continue
            
            info['selected'] = state
            values = list(self.tree.item(item, 'values'))
            if values:
                values[0] = checkmark
                self.tree.item(item, values=values)
            changed = True
        
        if changed:
            self._update_status()
            if self.on_selection_change:
                self.on_selection_change()
    
    def _invert_items_state(self, items):
        if not items:
            return
        
        changed = False
        for item in items:
            info = self._get_file_info_by_item(item)
            values = list(self.tree.item(item, 'values'))
            if not info or not values:
                continue
            
            info['selected'] = not info['selected']
            values[0] = "✓" if info['selected'] else "○"
            self.tree.item(item, values=values)
            changed = True
        
        if changed:
            self._update_status()
            if self.on_selection_change:
                self.on_selection_change()
    
    def _on_right_click(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            if item not in self.tree.selection():
                self.tree.selection_set(item)
            try:
                self.context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.context_menu.grab_release()
    
    def _on_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.tree.identify_column(event.x)
            item = self.tree.identify_row(event.y)
            
            if column == '#1' and item:
                info = self._get_file_info_by_item(item)
                if not info:
                    return
                
                target_state = not info['selected']
                items = self.tree.selection()
                if not items or item not in items:
                    items = (item,)
                
                self._set_items_state(items, target_state)
    
    def load_files(self, folder_path):
        if not os.path.exists(folder_path):
            return
        
        self.tree.delete(*self.tree.get_children())
        self.files_data = []
        self.file_index = {}
        
        pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.pdf')]
        
        for filename in pdf_files:
            filepath = os.path.join(folder_path, filename)
            if os.path.exists(filepath):
                try:
                    stat = os.stat(filepath)
                    size_kb = stat.st_size // 1024
                    mtime = dt.fromtimestamp(stat.st_mtime)
                    
                    file_info = {
                        'filename': filename,
                        'filepath': filepath,
                        'size_kb': size_kb,
                        'date': mtime.strftime("%d.%m.%Y"),
                        'time': mtime.strftime("%H:%M"),
                        'timestamp': mtime,
                        'selected': False
                    }
                    
                    self.files_data.append(file_info)
                except Exception as e:
                    print(f"Ошибка чтения файла {filename}: {e}")
        
        self.file_index = {info['filepath']: info for info in self.files_data}
        self.sort_by_column('date', reverse=True)
    
    def sort_by_column(self, column, reverse=None):
        if reverse is None:
            if self.sort_column == column:
                self.sort_reverse = not self.sort_reverse
            else:
                self.sort_column = column
                self.sort_reverse = (column == 'date')
        
        if column == 'name':
            self.files_data.sort(key=lambda x: x['filename'].lower(), reverse=self.sort_reverse)
        elif column == 'date':
            self.files_data.sort(key=lambda x: x['timestamp'], reverse=self.sort_reverse)
        elif column == 'size':
            self.files_data.sort(key=lambda x: x['size_kb'], reverse=self.sort_reverse)
        
        self._refresh_display()
    
    def _refresh_display(self):
        self.tree.delete(*self.tree.get_children())
        
        for file_info in self.files_data:
            checkmark = "✓" if file_info['selected'] else "○"
            self.tree.insert('', 'end', iid=file_info['filepath'], values=(
                checkmark,
                file_info['filename'],
                f"{file_info['size_kb']:,}",
                file_info['date'],
                file_info['time']
            ))
        
        self._update_status()
    
    def _update_status(self):
        total = len(self.files_data)
        selected = sum(1 for f in self.files_data if f['selected'])
        self.status_var.set(f"Файлов: {total} | Выбрано: {selected} ({selected/total*100:.1f}%)")
    
    def _apply_filter(self):
        search_text = self.search_var.get().lower()
        if not search_text:
            self._refresh_display()
            return
        
        self.tree.delete(*self.tree.get_children())
        
        for file_info in self.files_data:
            if search_text in file_info['filename'].lower():
                checkmark = "✓" if file_info['selected'] else "○"
                self.tree.insert('', 'end', iid=file_info['filepath'], values=(
                    checkmark,
                    file_info['filename'],
                    f"{file_info['size_kb']:,}",
                    file_info['date'],
                    file_info['time']
                ))
    
    def select_all(self):
        for file_info in self.files_data:
            file_info['selected'] = True
        self._refresh_display()
        if self.on_selection_change:
            self.on_selection_change()
    
    def select_none(self):
        for file_info in self.files_data:
            file_info['selected'] = False
        self._refresh_display()
        if self.on_selection_change:
            self.on_selection_change()
    
    def invert_selection(self):
        for file_info in self.files_data:
            file_info['selected'] = not file_info['selected']
        self._refresh_display()
        if self.on_selection_change:
            self.on_selection_change()
    
    def get_selected_files(self):
        return [f['filename'] for f in self.files_data if f['selected']]
    
    def get_selected_count(self):
        return sum(1 for f in self.files_data if f['selected'])


# ======================= ОСНОВНОЕ ОКНО =======================

class App(ttk.Frame):
    def __init__(self, master):
        super().__init__(master, padding=8)
        self.master = master
        self.master.title("КГС: обработчик PDF")
        self.master.geometry("1200x750")
        self.master.minsize(1000, 650)

        style = ttk.Style()
        try:
            style.theme_use("vista")
        except:
            pass

        self.folder_path = tk.StringVar(value="")
        self.points_folder = tk.StringVar(value="")
        self.move_folder_path = tk.StringVar(value="")
        self.status_text = tk.StringVar(value="")
        self.var_sort_points = tk.BooleanVar(value=True)

        self.selection_info_text = tk.StringVar(value="Выбрано: 0/0")
        self._tooltips = []

        try:
            style.configure("Cancel.TButton", foreground="#b00020")
        except Exception:
            pass

        # Прогресс и отмена
        self.progress_overall = tk.DoubleVar(value=0.0)
        self.progress_file_text = tk.StringVar(value="Файлы: 0/0")
        self.progress_page_text = tk.StringVar(value="Стр.: 0/0")
        self.current_file_text = tk.StringVar(value="")
        self.cancel_event = None
        self._progress_state = {
            "file_index": 0,
            "total_files": 0,
            "page_index": 0,
            "total_pages": 0,
            "filename": "",
        }

        self._build_toolbar()
        self._build_body_modern()
        self._build_bottom()

        self.processor = PDFProcessor(log_callback=self._append_log)
        self._update_selection_info()

    def _build_toolbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", pady=(0,6))
        ttk.Label(bar, text="Папка:").pack(side="left")
        self.entry_folder = ttk.Entry(bar, textvariable=self.folder_path)
        self.entry_folder.pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(bar, text="Обзор…", command=self.browse_folder).pack(side="left")
        ttk.Button(bar, text="Загрузить список…", command=self.load_files_from_txt).pack(side="left", padx=(6,0))

    def _build_body_modern(self):
        body = ttk.Panedwindow(self, orient="horizontal")
        body.pack(fill="both", expand=True)

        left = ttk.Frame(body)
        body.add(left, weight=3)
        self.file_selector = ModernFileSelector(left, on_selection_change=self._update_selection_info)
        self.file_selector.pack(fill='both', expand=True, padx=2, pady=2)

        right = ttk.Frame(body, padding=(8,0,0,0))
        body.add(right, weight=2)

        rec = ttk.LabelFrame(right, text="Распознавание")
        rec.pack(fill="x", pady=(0,8))

        self.var_import = tk.BooleanVar(value=False)
        self.var_debug = tk.BooleanVar(value=False)
        self.var_ignore_excel = tk.BooleanVar(value=False)
        self.var_move = tk.BooleanVar(value=False)

        cb_import = ttk.Checkbutton(rec, text="Импортировать координаты (TXT)", variable=self.var_import, command=self._toggle_points)
        cb_import.pack(anchor="w")
        cb_sort_points = ttk.Checkbutton(rec, text="Разложить каталоги по типам", variable=self.var_sort_points)
        cb_sort_points.pack(anchor="w", padx=(18,0))
        
        row_points = ttk.Frame(rec)
        row_points.pack(fill="x", pady=2)
        ttk.Label(row_points, text="Каталоги:").pack(side="left")
        self.entry_points = ttk.Entry(row_points, textvariable=self.points_folder, state="disabled")
        self.entry_points.pack(side="left", fill="x", expand=True, padx=6)
        self.btn_points = ttk.Button(row_points, text="…", width=3, command=self.browse_points, state="disabled")
        self.btn_points.pack(side="left")

        cb_debug = ttk.Checkbutton(rec, text="Режим отладки (сохранять OCR)", variable=self.var_debug)
        cb_debug.pack(anchor="w", pady=(4,0))

        btn_types = ttk.Button(rec, text="Типы коммуникаций…", command=self.open_comm_types_dialog)
        btn_types.pack(anchor="w", pady=(6,0))

        out = ttk.LabelFrame(right, text="После обработки")
        out.pack(fill="x")
        cb_ignore_excel = ttk.Checkbutton(out, text="Игнорировать запись в Excel", variable=self.var_ignore_excel)
        cb_ignore_excel.pack(anchor="w")
        cb_move = ttk.Checkbutton(out, text="Переместить обработанные файлы", variable=self.var_move, command=self._toggle_move)
        cb_move.pack(anchor="w", pady=(4,0))
        row_move = ttk.Frame(out)
        row_move.pack(fill="x", pady=2)
        ttk.Label(row_move, text="Папка:").pack(side="left")
        self.entry_move = ttk.Entry(row_move, textvariable=self.move_folder_path, state="disabled")
        self.entry_move.pack(side="left", fill="x", expand=True, padx=6)
        self.btn_move = ttk.Button(row_move, text="…", width=3, command=self.browse_move, state="disabled")
        self.btn_move.pack(side="left")

        self._tooltips.extend([
            Tooltip(cb_import, "Ищет и сохраняет таблицу координат точек в TXT."),
            Tooltip(cb_sort_points, "Складывает каталоги координат по подпапкам типа коммуникации."),
            Tooltip(cb_debug, "Сохраняет полный распознанный OCR-текст для каждого PDF."),
            Tooltip(btn_types, "Настройка ожидаемых типов коммуникаций."),
            Tooltip(cb_ignore_excel, "Не записывает результаты в Excel, только лог/координаты."),
            Tooltip(cb_move, "Перемещает обработанные PDF в указанную папку."),
        ])

    def _build_bottom(self):
        bottom = ttk.Frame(self)
        bottom.pack(fill="both", expand=False, pady=(6,0))

        runrow = ttk.Frame(bottom)
        runrow.pack(fill="x")
        left_actions = ttk.Frame(runrow)
        left_actions.pack(side="left")
        self.btn_run = ttk.Button(left_actions, text="Запустить обработку", command=self.run_processing)
        self.btn_run.pack(side="left")
        self.btn_cancel = ttk.Button(left_actions, text="Отмена", style="Cancel.TButton", state="disabled", command=self.cancel_processing)
        self.btn_cancel.pack(side="left", padx=(6,0))

        ttk.Label(runrow, textvariable=self.status_text, anchor="center").pack(side="left", fill="x", expand=True, padx=6)

        right_actions = ttk.Frame(runrow)
        right_actions.pack(side="right")
        self.btn_open_excel = ttk.Button(right_actions, text="Открыть Excel", state="disabled", command=self.open_excel)
        self.btn_open_excel.pack(side="right")

        progrow = ttk.Frame(bottom)
        progrow.pack(fill="x", pady=(4,0))
        self.progressbar = ttk.Progressbar(progrow, variable=self.progress_overall, maximum=100)
        self.progressbar.pack(side="left", fill="x", expand=True)
        ttk.Label(progrow, textvariable=self.progress_file_text, width=16, anchor="e").pack(side="left", padx=(6,0))

        detailrow = ttk.Frame(bottom)
        detailrow.pack(fill="x")
        ttk.Label(detailrow, textvariable=self.current_file_text).pack(side="left")
        ttk.Label(detailrow, textvariable=self.progress_page_text).pack(side="right")

        logf = ttk.LabelFrame(bottom, text="Лог")
        logf.pack(fill="both", expand=True, pady=(6,0))
        self.log = ScrolledText(logf, wrap="word", height=10, state="disabled")
        self.log.pack(fill="both", expand=True)

        footer = ttk.Frame(bottom)
        footer.pack(fill="x", pady=(4,0))
        # Ссылки TG и Repo в правом нижнем углу
        self.repo_label = tk.Label(
            footer,
            text="Repo",
            fg="#0088CC",
            cursor="hand2",
            font=("Segoe UI", 10, "underline"),
        )
        self.repo_label.pack(side="right", padx=(0, 8))
        self.repo_label.bind("<Button-1>", self.open_repo)

        self.telegram_label = tk.Label(
            footer,
            text="TG",
            fg="#0088CC",
            cursor="hand2",
            font=("Segoe UI", 10, "underline"),
        )
        self.telegram_label.pack(side="right", padx=(0, 8))
        self.telegram_label.bind("<Button-1>", self.open_telegram)

    def _append_log(self, msg):
        self.log.config(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.config(state="disabled")
        self.status_text.set(msg if len(msg) < 80 else msg[:77] + "…")

    def open_telegram(self, event=None):
        url = (TELEGRAM_URL or "").strip()
        if not url or "your_username" in url:
            messagebox.showinfo("Telegram", "Укажите ссылку на профиль в TELEGRAM_URL в коде.")
            return
        try:
            webbrowser.open(url)
        except Exception as e:
            messagebox.showerror("Telegram", f"Не удалось открыть ссылку: {e}")

    def open_repo(self, event=None):
        url = (REPO_URL or "").strip()
        if not url or "github.com" not in url:
            messagebox.showinfo("Repo", "Укажите ссылку на репозиторий в REPO_URL в коде.")
            return
        try:
            webbrowser.open(url)
        except Exception as e:
            messagebox.showerror("Repo", f"Не удалось открыть ссылку: {e}")

    def _reset_progress_ui(self, total_files=0):
        self._progress_state = {
            "file_index": 0,
            "total_files": total_files or 0,
            "page_index": 0,
            "total_pages": 0,
            "filename": "",
        }
        self.progress_overall.set(0.0)
        self.progress_file_text.set(f"Файлы: 0/{total_files}" if total_files else "Файлы: 0/0")
        self.progress_page_text.set("Стр.: 0/0")
        self.current_file_text.set("")

    def _on_progress(self, file_index=None, total_files=None, filename=None, page_index=None, total_pages=None, **_):
        def update():
            st = self._progress_state
            if total_files is not None:
                st["total_files"] = total_files
            if file_index is not None:
                st["file_index"] = file_index
            if total_pages is not None:
                st["total_pages"] = total_pages
            if page_index is not None:
                st["page_index"] = page_index
            if filename is not None:
                st["filename"] = filename

            tf = st.get("total_files") or 0
            fi = st.get("file_index") or 0
            tp = st.get("total_pages") or 0
            pi = st.get("page_index") or 0

            if tf:
                frac_in_file = (pi / tp) if tp else 0.0
                overall = ((max(fi - 1, 0) + frac_in_file) / tf) * 100.0
                self.progress_overall.set(max(0.0, min(overall, 100.0)))
                self.progress_file_text.set(f"Файлы: {fi}/{tf}")
            else:
                self.progress_overall.set(0.0)
                self.progress_file_text.set("Файлы: 0/0")

            self.progress_page_text.set(f"Стр.: {pi}/{tp}" if tp else "Стр.: 0/0")

            name = st.get("filename") or ""
            if name:
                self.current_file_text.set(name if len(name) <= 90 else name[:87] + "…")

        self.after(0, update)

    def cancel_processing(self):
        if self.cancel_event and not self.cancel_event.is_set():
            self.cancel_event.set()
            self.btn_cancel.config(state="disabled")
            self._append_log("Отмена… завершаю текущую страницу.")

    def _toggle_points(self):
        state = "normal" if self.var_import.get() else "disabled"
        self.entry_points.config(state=state)
        self.btn_points.config(state=state)
        if self.var_import.get() and not self.points_folder.get() and self.folder_path.get():
            self.points_folder.set(self.folder_path.get())

    def _toggle_move(self):
        state = "normal" if self.var_move.get() else "disabled"
        self.entry_move.config(state=state)
        self.btn_move.config(state=state)

    def _update_selection_info(self):
        total = len(getattr(self.file_selector, "files_data", []) or [])
        selected = self.file_selector.get_selected_count() if total else 0
        self.selection_info_text.set(f"Выбрано: {selected}/{total}")
        processing_active = str(self.btn_cancel.cget("state")) == "normal"
        if not processing_active:
            self.btn_run.config(state="normal" if selected else "disabled")

    def browse_folder(self):
        path = filedialog.askdirectory()
        if path:
            self.folder_path.set(path)
            self.file_selector.load_files(path)
            if not self.points_folder.get():
                self.points_folder.set(path)
            self._append_log(f"Загружена папка: {path}")
            self._update_selection_info()

    def browse_points(self):
        path = filedialog.askdirectory()
        if path:
            self.points_folder.set(path)

    def browse_move(self):
        path = filedialog.askdirectory()
        if path:
            if path == self.folder_path.get():
                messagebox.showwarning("Перемещение", "Папка перемещения не может совпадать с исходной.")
                return
            self.move_folder_path.set(path)

    def open_comm_types_dialog(self):
        dlg = tk.Toplevel(self.master)
        dlg.title("Типы коммуникаций")
        dlg.transient(self.master)
        dlg.grab_set()

        screen_w = dlg.winfo_screenwidth()
        screen_h = dlg.winfo_screenheight()
        w = min(560, max(360, screen_w - 80))
        h = min(420, max(300, screen_h - 120))
        dlg.geometry(f"{w}x{h}+{(screen_w - w)//2}+{(screen_h - h)//2}")
        dlg.minsize(min(520, w), min(360, h))

        dlg.rowconfigure(0, weight=1)
        dlg.columnconfigure(0, weight=1)

        main = ttk.Frame(dlg, padding=8)
        main.grid(row=0, column=0, sticky="nsew")
        main.rowconfigure(0, weight=1)
        main.columnconfigure(0, weight=1)

        types_local = self.processor.get_comm_types()
        types_local = [dict(name=t.get("name", ""), enabled=bool(t.get("enabled", True))) for t in types_local]

        CHECKED = "✓"
        UNCHECKED = ""

        tree_frame = ttk.Frame(main)
        tree_frame.grid(row=0, column=0, sticky="nsew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        columns = ("enabled", "name")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="extended")
        tree.heading("enabled", text="✓")
        tree.column("enabled", width=40, anchor="center", stretch=False)
        tree.heading("name", text="Тип коммуникации")
        tree.column("name", width=320, anchor="w", stretch=True)
        tree.grid(row=0, column=0, sticky="nsew")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=vsb.set)

        def render_rows(selected_names=None):
            selected_names = selected_names or set()
            tree.delete(*tree.get_children())
            for idx, t in enumerate(types_local):
                mark = CHECKED if t.get("enabled") else UNCHECKED
                tree.insert("", "end", iid=str(idx), values=(mark, t.get("name", "")))
                if t.get("name") in selected_names:
                    tree.selection_add(str(idx))

        def selected_items():
            return list(tree.selection())

        def set_items_enabled(items, enabled):
            for iid in items:
                try:
                    idx = int(iid)
                except ValueError:
                    continue
                if 0 <= idx < len(types_local):
                    types_local[idx]["enabled"] = enabled
                    tree.set(iid, "enabled", CHECKED if enabled else UNCHECKED)

        def invert_items(items):
            for iid in items:
                try:
                    idx = int(iid)
                except ValueError:
                    continue
                if 0 <= idx < len(types_local):
                    new_val = not bool(types_local[idx].get("enabled", True))
                    types_local[idx]["enabled"] = new_val
                    tree.set(iid, "enabled", CHECKED if new_val else UNCHECKED)

        def on_delete_selected():
            sel = selected_items()
            if not sel:
                return
            names = {types_local[int(iid)]["name"] for iid in sel if iid.isdigit() and int(iid) < len(types_local)}
            for iid in sorted((int(i) for i in sel if i.isdigit()), reverse=True):
                if 0 <= iid < len(types_local):
                    del types_local[iid]
            render_rows(selected_names=names & {t["name"] for t in types_local})

        def on_tree_click(event):
            if tree.identify("region", event.x, event.y) != "cell":
                return
            col = tree.identify_column(event.x)
            iid = tree.identify_row(event.y)
            if col == "#1" and iid:
                idx = int(iid)
                if 0 <= idx < len(types_local):
                    new_val = not bool(types_local[idx].get("enabled", True))
                    types_local[idx]["enabled"] = new_val
                    tree.set(iid, "enabled", CHECKED if new_val else UNCHECKED)
                    tree.selection_set(iid)
                    return "break"

        menu = tk.Menu(dlg, tearoff=0)
        menu.add_command(label="Включить", command=lambda: set_items_enabled(selected_items(), True))
        menu.add_command(label="Выключить", command=lambda: set_items_enabled(selected_items(), False))
        menu.add_command(label="Инвертировать", command=lambda: invert_items(selected_items()))
        menu.add_separator()
        menu.add_command(label="Удалить выбранные", command=on_delete_selected)

        def on_right_click(event):
            iid = tree.identify_row(event.y)
            if iid and iid not in tree.selection():
                tree.selection_set(iid)
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        tree.bind("<Button-1>", on_tree_click)
        tree.bind("<Button-3>", on_right_click)

        add_frame = ttk.Frame(main)
        add_frame.grid(row=1, column=0, sticky="ew", pady=(6, 2))
        add_frame.columnconfigure(1, weight=1)

        ttk.Label(add_frame, text="Новый тип:").grid(row=0, column=0, sticky="w")
        entry_new = ttk.Entry(add_frame)
        entry_new.grid(row=0, column=1, sticky="ew", padx=4)

        def on_add():
            name = entry_new.get().strip()
            if not name:
                return
            existing_lower = {t["name"].strip().lower() for t in types_local if t.get("name")}
            if name.lower() in existing_lower:
                messagebox.showinfo("Типы", "Такой тип уже есть в списке.")
                return
            types_local.append({"name": name, "enabled": True})
            render_rows(selected_names={name})
            entry_new.delete(0, "end")

        btn_add = ttk.Button(add_frame, text="Добавить", command=on_add)
        btn_add.grid(row=0, column=2, sticky="e")
        btn_del = ttk.Button(add_frame, text="Удалить выбранные", command=on_delete_selected)
        btn_del.grid(row=0, column=3, sticky="e", padx=(6, 0))

        entry_new.bind("<Return>", lambda _e: on_add())

        footer = ttk.Frame(main)
        footer.grid(row=2, column=0, sticky="ew", pady=(8, 0))

        def on_reset_defaults():
            nonlocal types_local
            types_local = [{"name": name, "enabled": True} for name in self.processor.default_comm_types]
            render_rows()

        def on_save():
            self.processor.update_comm_types(types_local)
            self._append_log("Обновлён список типов коммуникаций.")
            dlg.destroy()

        btn_save = ttk.Button(footer, text="Сохранить", command=on_save)
        btn_cancel = ttk.Button(footer, text="Отмена", command=dlg.destroy)
        btn_defaults = ttk.Button(footer, text="По умолчанию", command=on_reset_defaults)

        btn_save.pack(side="right")
        btn_cancel.pack(side="right", padx=(6, 0))
        btn_defaults.pack(side="right", padx=(0, 6))

        render_rows()

    def load_files_from_txt(self):
        if not self.folder_path.get():
            messagebox.showwarning("Загрузка списка", "Сначала выберите папку с PDF.")
            return
        fp = filedialog.askopenfilename(title="TXT со списком", filetypes=[("TXT", "*.txt")])
        if not fp:
            return
        try:
            with open(fp, 'r', encoding='utf-8') as f:
                names = {line.strip() for line in f if line.strip()}
            self.file_selector.load_files(self.folder_path.get())
            for f in self.file_selector.files_data:
                f['selected'] = f['filename'] in names
            self.file_selector._refresh_display()
            self._append_log(f"Загружен список: {len(names)} файлов")
            self._update_selection_info()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить список: {e}")

    def run_processing(self):
        folder = self.folder_path.get().strip()
        if not folder:
            messagebox.showwarning("Ошибка", "Выберите папку.")
            return
        selected = self.file_selector.get_selected_files()
        if not selected:
            messagebox.showwarning("Ошибка", "Выберите хотя бы один PDF-файл.")
            return

        self.cancel_event = threading.Event()
        self.processor.set_cancel_event(self.cancel_event)
        self.processor.set_progress_callback(self._on_progress)
        self._reset_progress_ui(total_files=len(selected))
        self.btn_run.config(state="disabled")
        self.btn_cancel.config(state="normal")
        self.status_text.set("Работаю…")
        threading.Thread(target=self._process_files_thread, args=(folder, selected), daemon=True).start()

    def _process_files_thread(self, folder, selected_files):
        self.after(0, lambda: self.log.config(state="normal"))
        self.after(0, lambda: self.log.delete("1.0", "end"))
        self.after(0, lambda: self.log.config(state="disabled"))
        self.after(0, lambda: self.btn_open_excel.config(state="disabled"))
        self.after(0, lambda: self.status_text.set("Работаю…"))

        self.processor.problem_files = []
        self.processor.field_stats = defaultdict(int)
        self.processor.import_points = self.var_import.get()
        self.processor.sort_points_by_comm = self.var_sort_points.get()
        self.processor.debug_mode = self.var_debug.get()
        self.processor.ignore_excel = self.var_ignore_excel.get()
        self.processor.points_folder = self.points_folder.get().strip()

        error_happened = False
        try:
            path = self.processor.process_selected_files(
                folder, selected_files,
                self.move_folder_path.get().strip() if self.var_move.get() else None
            )
            if path and os.path.exists(path):
                self.after(0, lambda: self.btn_open_excel.config(state="normal"))
                self.after(0, lambda: self._append_log(f"✓ Excel создан: {os.path.basename(path)}"))
            elif self.processor.ignore_excel:
                self.after(0, lambda: self._append_log("ℹ Excel отключён."))
            else:
                self.after(0, lambda: self._append_log("⚠ Excel не создан — проверьте логи."))
        except Exception as e:
            error_happened = True
            self.after(0, lambda: self._append_log(f"❌ Ошибка: {e}"))
            self.after(0, lambda: messagebox.showerror("Ошибка", str(e)))

        self.after(0, lambda: self.file_selector.load_files(folder))
        self.after(0, self._update_selection_info)
        was_cancelled = bool(self.cancel_event and self.cancel_event.is_set()) or getattr(self.processor, "cancelled", False)
        if was_cancelled:
            self.after(0, lambda: self.status_text.set("Отменено"))
        else:
            self.after(0, lambda: self.status_text.set("Готово"))
            if not error_happened:
                self.after(0, lambda: self.progress_overall.set(100.0))
        self.after(0, lambda: self.btn_run.config(state="normal"))
        self.after(0, lambda: self.btn_cancel.config(state="disabled"))

    def open_excel(self):
        path = self.processor.output_excel_path
        if path and os.path.exists(path):
            try:
                os.startfile(path)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось открыть Excel: {e}")
        else:
            messagebox.showerror("Ошибка", "Файл Excel не найден.")


# ======================= ЗАПУСК =======================

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    app.pack(fill="both", expand=True)
    root.mainloop()
